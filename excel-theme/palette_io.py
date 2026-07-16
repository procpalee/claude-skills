# -*- coding: utf-8 -*-
"""
팔레트 ↔ 엑셀 왕복 도구 (Export / Import).

  export :  현재 excel_theme.py 의 설정을 '편집용 엑셀'로 내보낸다.
            → 사용자가 엑셀에서 색(채우기)·폰트·표옵션·크기·숫자서식을 바꿔 저장.
  import :  그 엑셀을 읽어 excel_theme.py 를 자동 반영한다(.bak 백업, 바뀐 항목만).

사용:
  python palette_io.py export  <out.xlsx>
  python palette_io.py import  <in.xlsx>

엑셀 규칙:
  • 색(colors)        = C열 "채우기색" (엑셀 채우기 버킷)
  • 폰트/표옵션/크기  = C열 "값(텍스트/숫자)"
  • 시트 = 테마(valuation/closing/general), 그리고 공통 "숫자서식" 시트 1개.
"""
import os
import re
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_theme.py")

COLOR_ROLES = ["primary", "primary_dark", "secondary", "secondary_lt",
               "band", "subheader", "surface", "text", "note", "accent",
               "negative", "header_font", "title_color", "border_in", "border_out"]
FONT_KEYS = ["font_name", "font_excel", "font_fallback"]
FLAG_KEYS = ["excel_table_style", "excel_zebra", "excel_neg_red", "excel_highlight_neg"]
SIZE_KEYS = ["excel_title_size", "excel_size_header", "excel_size_body",
             "excel_row_header", "excel_row_body"]
NF_SHEET = "숫자서식"


def _load_ns():
    """excel_theme.py 소스를 매번 새로 실행해 네임스페이스를 얻는다 (.pyc 캐시 회피)."""
    ns = {}
    with open(SRC, encoding="utf-8") as f:
        exec(compile(f.read(), SRC, "exec"), ns)
    return ns


# ──────────────────────────────────────────────────────────────
# EXPORT
# ──────────────────────────────────────────────────────────────
def export(out_path):
    ns = _load_ns()
    palettes = ns["PALETTES"]
    number_formats = ns["NUMBER_FORMATS"]

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    sec = Font(bold=True, italic=True)

    for theme, pal in palettes.items():
        ws = wb.create_sheet(theme[:31])
        ws["B2"] = f"팔레트 편집 — {theme}  (C열을 바꿔 저장 → import)"
        ws["B2"].font = Font(bold=True, size=12)
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 16
        r = 4
        ws.cell(r, 2, "항목").font = bold
        ws.cell(r, 3, "여기를 바꾸세요").font = bold
        ws.cell(r, 4, "현재값(참고)").font = bold
        r += 1
        # 색
        ws.cell(r, 2, "── 색 (셀 채우기색을 바꾸세요) ──").font = sec
        r += 1
        for role in COLOR_ROLES:
            hexv = pal["colors"].get(role)
            if not hexv:
                continue
            ws.cell(r, 2, role).border = border
            sw = ws.cell(r, 3, "")
            sw.fill = PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")
            sw.border = border
            ws.cell(r, 4, f"#{hexv}").border = border
            r += 1
        # 폰트
        r += 1
        ws.cell(r, 2, "── 폰트 (값=폰트이름) ──").font = sec
        r += 1
        for k in FONT_KEYS:
            if k not in pal:
                continue
            ws.cell(r, 2, k).border = border
            c = ws.cell(r, 3, pal[k])
            try:
                c.font = Font(name=pal[k])
            except Exception:
                pass
            c.border = border
            r += 1
        # 표 옵션
        r += 1
        ws.cell(r, 2, "── 엑셀 표 옵션 (값) ──").font = sec
        r += 1
        flag_hint = {"excel_table_style": "rules 또는 grid",
                     "excel_zebra": "True/False", "excel_neg_red": "True/False",
                     "excel_highlight_neg": "True/False"}
        for k in FLAG_KEYS:
            if k not in pal:
                continue
            ws.cell(r, 2, k).border = border
            ws.cell(r, 3, str(pal[k])).border = border
            ws.cell(r, 4, flag_hint.get(k, "")).border = border
            r += 1
        # 크기/행높이
        r += 1
        ws.cell(r, 2, "── 크기/행높이 (값=숫자) ──").font = sec
        r += 1
        size_hint = {"excel_title_size": "제목 pt", "excel_size_header": "헤더 pt",
                     "excel_size_body": "본문 pt", "excel_row_header": "헤더 행높이",
                     "excel_row_body": "본문 행높이"}
        for k in SIZE_KEYS:
            if k not in pal:
                continue
            ws.cell(r, 2, k).border = border
            ws.cell(r, 3, pal[k]).border = border
            ws.cell(r, 4, size_hint.get(k, "")).border = border
            r += 1

    # 공통 숫자서식 시트
    nfws = wb.create_sheet(NF_SHEET)
    nfws["B2"] = "공통 숫자서식 (C열 서식 문자열을 바꿔 저장)"
    nfws["B2"].font = Font(bold=True, size=12)
    nfws.column_dimensions["B"].width = 16
    nfws.column_dimensions["C"].width = 42
    nfws.cell(4, 2, "키").font = bold
    nfws.cell(4, 3, "서식 문자열").font = bold
    rr = 5
    for k, v in number_formats.items():
        nfws.cell(rr, 2, k).border = border
        nfws.cell(rr, 3, v).border = border
        rr += 1

    wb.save(out_path)
    print("내보냄:", out_path, f"(테마 {len(palettes)}개 + 숫자서식)")


# ──────────────────────────────────────────────────────────────
# IMPORT
# ──────────────────────────────────────────────────────────────
def _read_xlsx(path):
    wb = load_workbook(path)
    themes = {}
    number_formats = None
    for ws in wb.worksheets:
        if ws.title == NF_SHEET:
            number_formats = {}
            for row in ws.iter_rows(min_col=2, max_col=3):
                k, v = row[0].value, row[1].value
                if isinstance(k, str) and k.strip() and v is not None:
                    number_formats[k.strip()] = str(v)
            continue
        colors, fonts, flags, sizes = {}, {}, {}, {}
        for row in ws.iter_rows(min_col=2, max_col=3):
            label, cell = row[0].value, row[1]
            if not isinstance(label, str):
                continue
            label = label.strip()
            if label in COLOR_ROLES:
                f = cell.fill
                if f and f.patternType == "solid" and isinstance(f.fgColor.rgb, str):
                    colors[label] = f.fgColor.rgb[-6:].upper()
            elif label in FONT_KEYS:
                if cell.value:
                    fonts[label] = str(cell.value).strip()
            elif label in FLAG_KEYS:
                if cell.value is not None:
                    flags[label] = str(cell.value).strip()
            elif label in SIZE_KEYS:
                try:
                    sizes[label] = int(float(cell.value))
                except (TypeError, ValueError):
                    pass
        themes[ws.title] = {"colors": colors, "fonts": fonts,
                            "flags": flags, "sizes": sizes}
    return themes, number_formats


def _theme_block_span(src, theme):
    m = re.search(r'"%s"\s*:\s*\{' % re.escape(theme), src)
    if not m:
        return None
    start = src.index("{", m.start())
    depth = 0
    for j in range(start, len(src)):
        if src[j] == "{":
            depth += 1
        elif src[j] == "}":
            depth -= 1
            if depth == 0:
                return (start, j + 1)
    return None


def _apply_updates(block, colors, fonts, flags, sizes):
    changed = []
    for role, hexv in colors.items():
        pat = re.compile(r'("%s"\s*:\s*")[0-9A-Fa-f]{6}(")' % re.escape(role))
        new, n = pat.subn(lambda m: m.group(1) + hexv + m.group(2), block)
        if n and new != block:
            block = new; changed.append(f"{role}=#{hexv}")
    for k, val in fonts.items():
        pat = re.compile(r'("%s"\s*:\s*")[^"]*(")' % re.escape(k))
        new, n = pat.subn(lambda m: m.group(1) + val + m.group(2), block)
        if n and new != block:
            block = new; changed.append(f"{k}={val}")
    for k, val in sizes.items():
        pat = re.compile(r'("%s"\s*:\s*)\d+' % re.escape(k))
        new, n = pat.subn(lambda m: m.group(1) + str(int(val)), block)
        if n and new != block:
            block = new; changed.append(f"{k}={val}")
    for k, val in flags.items():
        if k == "excel_table_style":
            pat = re.compile(r'("%s"\s*:\s*")[^"]*(")' % re.escape(k))
            new, n = pat.subn(lambda m: m.group(1) + val + m.group(2), block)
        else:
            b = "True" if str(val).strip().lower() in ("true", "1", "on", "yes") else "False"
            pat = re.compile(r'("%s"\s*:\s*)(True|False)' % re.escape(k))
            new, n = pat.subn(lambda m: m.group(1) + b, block)
            val = b
        if n and new != block:
            block = new; changed.append(f"{k}={val}")
    return block, changed


def _nf_literal(val):
    """숫자서식 문자열 → 파이썬 리터럴(내용에 맞춰 따옴표 선택)."""
    if '"' in val and "'" not in val:
        return "'" + val + "'"
    if '"' in val and "'" in val:
        return repr(val)
    return '"' + val + '"'


def _replace_nf(src, key, val):
    pat = re.compile(r'("%s"\s*:\s*)(?:"(?:[^"\\]|\\.)*"|\'(?:[^\'\\]|\\.)*\')'
                     % re.escape(key))
    return pat.subn(lambda m: m.group(1) + _nf_literal(val), src)


def import_(in_path):
    ns = _load_ns()
    palettes = ns["PALETTES"]
    cur_nf = ns["NUMBER_FORMATS"]
    themes, number_formats = _read_xlsx(in_path)

    with open(SRC, encoding="utf-8") as f:
        src = f.read()
    total = []

    for theme, upd in themes.items():
        if theme not in palettes:
            print(f"  (건너뜀: 알 수 없는 테마 시트 '{theme}')")
            continue
        span = _theme_block_span(src, theme)
        if not span:
            print(f"  (건너뜀: 소스에서 '{theme}' 블록을 못 찾음)")
            continue
        s, e = span
        new_block, changed = _apply_updates(src[s:e], upd["colors"], upd["fonts"],
                                            upd["flags"], upd["sizes"])
        if changed:
            src = src[:s] + new_block + src[e:]
            total.append((theme, changed))

    # 숫자서식 (공통)
    if number_formats:
        nf_changed = []
        for k, v in number_formats.items():
            if k in cur_nf and v != cur_nf[k]:
                src, n = _replace_nf(src, k, v)
                if n:
                    nf_changed.append(k)
        if nf_changed:
            total.append((NF_SHEET, nf_changed))

    if not total:
        print("변경 사항 없음.")
        return
    shutil.copyfile(SRC, SRC + ".bak")
    with open(SRC, "w", encoding="utf-8") as f:
        f.write(src)
    print("excel_theme.py 갱신 완료 (백업: excel_theme.py.bak)")
    for name, ch in total:
        print(f"  [{name}] " + ", ".join(ch))


# ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3 or sys.argv[1] not in ("export", "import"):
        print("사용법: python palette_io.py export <out.xlsx> | import <in.xlsx>")
        sys.exit(2)
    if sys.argv[1] == "export":
        export(sys.argv[2])
    else:
        import_(sys.argv[2])
