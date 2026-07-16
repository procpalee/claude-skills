# -*- coding: utf-8 -*-
"""
엑셀 테마 샘플 생성기.

table_samples.py 의 SAMPLES(여러 표 유형)를 읽어, 테마마다
한 워크북(샘플마다 한 시트)으로 미리보기를 만든다.
    python build_preview.py
  → preview_<theme>.xlsx  (시트: 01_단순표, 02_합계만, ...)

표 규칙(소계/합계/섹션/숫자서식/줄무늬 등)을 바꾸려면 table_samples.py 를 편집한다.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from excel_theme import (apply_theme, style_header_row, style_total_row,
                         style_subheader_row, mark_cells, highlight_threshold,
                         get_theme, THEMES)
from table_samples import SAMPLES


def build_header_demo(ws, theme_name):
    """헤더 3색(제목 B2 / 본문 기본 / 본문 세컨더리)을 한 시트에 보여준다."""
    th = get_theme(theme_name)
    ws["B2"] = "헤더 3종 — 제목 / 본문 기본 / 본문 세컨더리"
    # 표1: 본문 기본 헤더
    for j, h in enumerate(["계정", "당기", "전기"]):
        ws.cell(4, 2 + j, h)
    for i, row in enumerate([["매출액", 15000, 12000], ["매출원가", -9500, -8000],
                             ["영업이익", 2500, 1900]]):
        for j, v in enumerate(row):
            ws.cell(5 + i, 2 + j, v)
    apply_theme(ws, theme=theme_name, header_row=4, data_range="B4:D7", title_cell="B2",
                number_format_cols={"C": "accounting", "D": "accounting"}, autofit=False)
    # 표2: 본문 세컨더리 헤더 (한 시트에 2차로 나오는 표)
    cap = ws["B9"]
    cap.value = "▸ 하위 보조 표 (세컨더리 헤더)"
    cap.font = Font(name=th["font_name"], size=th["font_size_body"], bold=True,
                    color=th.get("text_color"))
    for j, h in enumerate(["조정 항목", "금액"]):
        ws.cell(10, 2 + j, h)
    for i, row in enumerate([["가산조정", 300], ["차감조정", -120], ["조정계", 180]]):
        for j, v in enumerate(row):
            ws.cell(11 + i, 2 + j, v)
    apply_theme(ws, theme=theme_name, header_row=10, data_range="B10:C13", title_cell=None,
                number_format_cols={"C": "accounting"}, autofit=False, margin=False)
    style_header_row(ws, th, header_row=10, min_col=2, max_col=3, secondary=True)
    for col, w in {"B": 16, "C": 13, "D": 13}.items():
        ws.column_dimensions[col].width = w


def build_sheet(ws, spec, theme_name):
    th = get_theme(theme_name)
    ws["B2"] = spec.get("title", spec["name"])

    headers = spec["headers"]
    n_cols = len(headers)
    last_col = get_column_letter(1 + n_cols)   # B부터 시작하므로 +1

    group_header = spec.get("group_header")
    if group_header:
        # 4행: 묶음헤더(병합), 5행: 열 헤더
        col = 2
        for label, span in group_header:
            ws.cell(row=4, column=col, value=label)
            if span > 1:
                ws.merge_cells(start_row=4, start_column=col,
                               end_row=4, end_column=col + span - 1)
            col += span
        header_row = 5
        range_top = 4            # 묶음헤더까지 테두리에 포함
    else:
        header_row = 4
        range_top = 4

    for j, h in enumerate(headers, start=2):
        ws.cell(row=header_row, column=j, value=h)

    data_start = header_row + 1
    for i, row in enumerate(spec["rows"]):
        for j, val in enumerate(row, start=2):
            if val is not None:
                ws.cell(row=data_start + i, column=j, value=val)
    last_row = data_start + len(spec["rows"]) - 1

    # 줄무늬/음수강조 기본값은 테마(팔레트)가 정함. 샘플에서 명시하면 그 값 우선.
    zebra = spec.get("zebra", th.get("excel_zebra", True))
    hl_neg = spec.get("highlight_neg", th.get("excel_highlight_neg", True))
    apply_theme(
        ws, theme=theme_name, header_row=header_row,
        data_range=f"B{range_top}:{last_col}{last_row}",
        title_cell="B2",
        number_format_cols=spec.get("number_format_cols"),
        zebra=zebra,
        freeze=spec.get("freeze", False),   # 기본 규칙: 틀 고정 안 함
        highlight_neg=hl_neg,
    )

    # 묶음헤더 행: 본문 헤더와 같은 색 + (frame의) 얇은 테두리로만 그룹 구분
    if group_header:
        style_header_row(ws, th, header_row=4, min_col=2, max_col=1 + n_cols)

    # rows 기준 인덱스를 실제 시트 행으로 변환해 규칙 적용
    def sheet_row(idx):
        return data_start + idx - 1

    for idx in spec.get("subheader_rows", []):
        style_subheader_row(ws, th, row=sheet_row(idx), min_col=2, max_col=1 + n_cols)
    for idx in spec.get("subtotal_rows", []):
        style_total_row(ws, th, row=sheet_row(idx), min_col=2, max_col=1 + n_cols, fill=False)
    for idx in spec.get("total_rows", []):
        style_total_row(ws, th, row=sheet_row(idx), min_col=2, max_col=1 + n_cols, fill=True)

    # 입력(크림)/타시트 연결(파랑) 셀 표시
    if spec.get("input_cells") or spec.get("linked_cells"):
        mark_cells(ws, th, input=spec.get("input_cells", []),
                   linked=spec.get("linked_cells", []))

    # 조건부 서식(임계값 강조)
    ht = spec.get("highlight_threshold")
    if ht:
        highlight_threshold(ws, th, ht["range"], ht["operator"], ht["value"])


def build_one(theme_name, out_path):
    wb = Workbook()
    wb.remove(wb.active)   # 기본 시트 제거
    for spec in SAMPLES:
        ws = wb.create_sheet(spec["name"][:31])
        build_sheet(ws, spec, theme_name)
    build_header_demo(wb.create_sheet("헤더3종_데모"), theme_name)
    wb.save(out_path)
    print(f"  생성: {out_path}  (시트 {len(SAMPLES)}개)")


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    print("엑셀 샘플 생성 중...")
    for name, pal in THEMES.items():
        if pal.get("legacy"):        # legacy 팔레트(navy/charcoal)는 미리보기 생략
            continue
        build_one(name, os.path.join(here, f"preview_{name}.xlsx"))
    print("완료.")


if __name__ == "__main__":
    main()
