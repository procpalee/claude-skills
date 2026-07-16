# -*- coding: utf-8 -*-
"""
실무 파일(조서·정산표 등) 분석기 — 테마/표 규칙 개선의 '결정 재료'를 만든다.

    python analyze_reference.py <실무.xlsx> [출력폴더]
  → <출력폴더>\<이름>.analysis.md

보고서 내용 (시트별 + 워크북 요약):
  - 구조: 사용 영역, 병합 셀, 틀 고정, 열너비 분포
  - 스타일 센서스: (채우기·폰트·테두리·숫자서식) 고유 조합 빈도 + 대표 셀
                   + 역할 휴리스틱(헤더/합계/입력셀 후보) 태깅
  - 수식 센서스: 수식/상수 비율, 함수 빈도, 타시트 참조, 에러 셀
  - 색·숫자서식 빈도를 excel_theme.py 의 팔레트/NUMBER_FORMATS 와 대조
    → 테마에 없는 항목을 "추가 후보"로 표시

분석기는 결정하지 않는다 — 보고서를 Claude 에게 주고
"이 패턴을 closing 테마/샘플에 반영해줘" 처럼 다음 행동을 지시하는 용도.
"""
import os
import re
import sys
from collections import Counter, defaultdict

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, _ROOT)
sys.path.insert(0, _HERE)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_theme import NUMBER_FORMATS, PALETTES
from style_capture import _norm_nf, build_theme_resolver, resolve_color

MAX_CELLS_PER_SHEET = 40_000     # 초대형 시트 보호 (넘으면 앞부분만)
ERROR_VALUES = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
FUNC_RE = re.compile(r"([A-Z가-힣_][A-Z0-9가-힣_.]*)\s*\(", re.IGNORECASE)


# ── 팔레트 역인덱스: hex → "테마.토큰" 목록 ───────────────────────────
def _palette_index():
    idx = defaultdict(list)
    for theme, pal in PALETTES.items():
        for token, hexv in pal["colors"].items():
            idx[str(hexv).upper()].append(f"{theme}.{token}")
    return idx


def _nf_index():
    return {_norm_nf(v): k for k, v in NUMBER_FORMATS.items()}


def _is_cream(hex6):
    """크림 계열(입력셀 후보, FFFBEF 등): 밝은 노랑끼."""
    try:
        r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
    except (TypeError, ValueError):
        return False
    return r >= 240 and g >= 235 and b <= 245 and (r + g) / 2 - b >= 8


def _is_blue_tint(hex6):
    """옅은 파랑(연결셀 후보, CCECFF 등)."""
    try:
        r, g, b = int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)
    except (TypeError, ValueError):
        return False
    return b >= 230 and b - r >= 20


def _border_sig(border):
    """테두리 시그니처: 'T:medium B:thin' 식 요약."""
    parts = []
    for label, side in (("T", border.top), ("B", border.bottom),
                        ("L", border.left), ("R", border.right)):
        if side is not None and side.style:
            parts.append(f"{label}:{side.style}")
    return " ".join(parts) or "-"


def _role_hint(fill, bold, border_sig, nf, has_number):
    hints = []
    if fill and _is_cream(fill):
        hints.append("입력셀 후보")
    if fill and _is_blue_tint(fill):
        hints.append("연결셀 후보")
    if bold and fill and not has_number:
        hints.append("헤더/섹션 후보")
    if bold and ("T:medium" in border_sig or "T:thin" in border_sig) and has_number:
        hints.append("소계/합계 후보")
    if "T:medium" in border_sig and "B:medium" in border_sig:
        hints.append("statement 상하단선")
    return ", ".join(hints)


def analyze_sheet(ws_f, ws_v, resolver):
    """ws_f: 수식 보기 / ws_v: 캐시값 보기. 시트 1개 분석 dict."""
    info = {"name": ws_f.title, "dim": ws_f.calculate_dimension(),
            "max_row": ws_f.max_row, "max_col": ws_f.max_column}
    info["merged"] = [str(r) for r in list(ws_f.merged_cells.ranges)[:8]]
    info["merged_count"] = len(ws_f.merged_cells.ranges)
    info["freeze"] = str(ws_f.freeze_panes) if ws_f.freeze_panes else None

    widths = []
    for d in ws_f.column_dimensions.values():
        if d.width:
            n = ((d.max or 0) - (d.min or 0) + 1) if d.min else 1
            widths.extend([round(d.width, 1)] * min(n, 50))
    info["width_freq"] = Counter(widths).most_common(6)

    style_census = Counter()
    style_repr = {}
    fill_freq, fontc_freq, font_freq, nf_freq = Counter(), Counter(), Counter(), Counter()
    n_cells = n_formula = n_const = n_error = 0
    func_freq = Counter()
    xrefs = Counter()
    errors = []

    scanned = 0
    for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
        if scanned > MAX_CELLS_PER_SHEET:
            info["truncated"] = True
            break
        for cf, cv in zip(row_f, row_v):
            scanned += 1
            v = cf.value
            has_style = (cf.has_style if hasattr(cf, "has_style") else True)
            if v is None and not has_style:
                continue
            n_cells += 1

            # 수식/상수
            if isinstance(v, str) and v.startswith("="):
                n_formula += 1
                for fn in FUNC_RE.findall(v):
                    if not fn.startswith("_xl"):
                        func_freq[fn.upper()] += 1
                for m in re.findall(r"'?\[?([^\[\]'!=,()]+?)'?!", v):
                    if m != ws_f.title:
                        xrefs[m] += 1
            elif v is not None:
                n_const += 1
            ev = cv.value
            if isinstance(ev, str) and ev in ERROR_VALUES:
                n_error += 1
                if len(errors) < 10:
                    errors.append(f"{cf.coordinate}: {ev}")

            if not has_style and v is None:
                continue
            # 스타일 센서스
            fill = None
            if cf.fill is not None and cf.fill.patternType == "solid":
                fill = resolve_color(cf.fill.fgColor, resolver)
            font = cf.font
            fname = font.name or "(기본)"
            fcolor = resolve_color(font.color, resolver) or "000000"
            bsig = _border_sig(cf.border)
            nf = _norm_nf(cf.number_format)
            if fill:
                fill_freq[fill] += 1
            if fcolor != "000000":
                fontc_freq[fcolor] += 1
            font_freq[(fname, float(font.size or 0), bool(font.bold))] += 1
            if nf != "General":
                nf_freq[nf] += 1
            key = (fill, fname, float(font.size or 0), bool(font.bold),
                   fcolor, bsig, nf)
            style_census[key] += 1
            if key not in style_repr:
                style_repr[key] = cf.coordinate

    info.update(n_cells=n_cells, n_formula=n_formula, n_const=n_const,
                n_error=n_error, errors=errors,
                func_freq=func_freq.most_common(10),
                xrefs=xrefs.most_common(8),
                style_census=style_census.most_common(12),
                style_repr=style_repr,
                fill_freq=fill_freq, fontc_freq=fontc_freq,
                font_freq=font_freq.most_common(6),
                nf_freq=nf_freq)
    return info


def render_report(xlsx_path, sheets_info, out_md):
    pal_idx = _palette_index()
    nf_idx = _nf_index()
    L = [f"# 실무 파일 분석 — {os.path.basename(xlsx_path)}", ""]
    L.append(f"원본: `{xlsx_path}`")
    L.append(f"시트 {len(sheets_info)}개: " + ", ".join(s["name"] for s in sheets_info))
    L.append("")

    # ── 워크북 요약: 색/서식/폰트 vs 테마 ──
    all_fill, all_fontc, all_nf = Counter(), Counter(), Counter()
    for s in sheets_info:
        all_fill.update(s["fill_freq"])
        all_fontc.update(s["fontc_freq"])
        all_nf.update(s["nf_freq"])

    L.append("## 워크북 요약 — 테마와 대조")
    L.append("")
    L.append("### 채우기색 빈도 (상위 12)")
    L.append("")
    L.append("| 색 | 사용 횟수 | 팔레트 매칭 |")
    L.append("|---|---|---|")
    for hexv, cnt in all_fill.most_common(12):
        match = ", ".join(pal_idx.get(hexv, [])) or "**(팔레트에 없음 — 추가 후보)**"
        L.append(f"| `#{hexv}` | {cnt} | {match} |")
    L.append("")
    L.append("### 글자색 빈도 (검정 제외, 상위 8)")
    L.append("")
    L.append("| 색 | 사용 횟수 | 팔레트 매칭 |")
    L.append("|---|---|---|")
    for hexv, cnt in all_fontc.most_common(8):
        match = ", ".join(pal_idx.get(hexv, [])) or "**(팔레트에 없음 — 추가 후보)**"
        L.append(f"| `#{hexv}` | {cnt} | {match} |")
    L.append("")
    L.append("### 숫자서식 빈도 (상위 12)")
    L.append("")
    L.append("| 서식 | 사용 횟수 | NUMBER_FORMATS 매칭 |")
    L.append("|---|---|---|")
    for nf, cnt in all_nf.most_common(12):
        key = nf_idx.get(nf)
        match = f"`{key}`" if key else "**(미보유 — 추가 후보)**"
        L.append(f"| `{nf}` | {cnt} | {match} |")
    L.append("")

    # ── 시트별 ──
    for s in sheets_info:
        L.append(f"## 시트: {s['name']}")
        L.append("")
        L.append(f"- 사용 영역: `{s['dim']}` ({s['max_row']}행 × {s['max_col']}열)"
                 + (" — **일부만 스캔**" if s.get("truncated") else ""))
        L.append(f"- 병합 셀: {s['merged_count']}개"
                 + (f" (예: {', '.join(s['merged'][:5])})" if s["merged"] else ""))
        L.append(f"- 틀 고정: {s['freeze'] or '없음'}")
        if s["width_freq"]:
            L.append("- 열너비 분포: " + ", ".join(f"{w}({c}열)" for w, c in s["width_freq"]))
        total_vals = s["n_formula"] + s["n_const"]
        if total_vals:
            pct = 100.0 * s["n_formula"] / total_vals
            L.append(f"- 수식 {s['n_formula']} / 상수 {s['n_const']} (수식 비율 {pct:.0f}%)")
        if s["func_freq"]:
            L.append("- 함수 빈도: " + ", ".join(f"{f}×{c}" for f, c in s["func_freq"]))
        if s["xrefs"]:
            L.append("- 타시트 참조: " + ", ".join(f"{n}×{c}" for n, c in s["xrefs"]))
        if s["n_error"]:
            L.append(f"- **에러 셀 {s['n_error']}개**: " + "; ".join(s["errors"]))
        L.append("")
        L.append("| 대표 셀 | 채우기 | 폰트 | 크기 | 굵게 | 글자색 | 테두리 | 숫자서식 | 셀 수 | 역할 추정 |")
        L.append("|---|---|---|---|---|---|---|---|---|---|")
        for key, cnt in s["style_census"]:
            fill, fname, fsize, fbold, fcolor, bsig, nf = key
            hint = _role_hint(fill, fbold, bsig, nf, nf != "General")
            L.append(f"| {s['style_repr'][key]} | {('#' + fill) if fill else '-'} "
                     f"| {fname} | {fsize:g} | {'O' if fbold else ''} "
                     f"| {('#' + fcolor) if fcolor != '000000' else '-'} "
                     f"| {bsig} | `{nf}` | {cnt} | {hint} |")
        L.append("")

    L.append("---")
    L.append("다음 행동 예시 — 이 보고서를 Claude 에게 주고:")
    L.append("- \"팔레트에 없는 색 #XXXXXX 를 closing 테마 ○○ 토큰으로 추가해줘\"")
    L.append("- \"미보유 숫자서식을 NUMBER_FORMATS 에 추가해줘\"")
    L.append("- \"이 표 구조(섹션/소계/합계 패턴)를 tools/sample_specs.py 샘플로 만들어줘\"")
    L.append("  → 반영 후  dev.bat -Samples  →  엑셀 검수  →  dev.bat -Diff  로 검증")

    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


def analyze(xlsx_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    print(f"분석 중: {xlsx_path}")
    wb_f = load_workbook(xlsx_path, data_only=False)
    wb_v = load_workbook(xlsx_path, data_only=True)
    resolver = build_theme_resolver(wb_f)
    sheets_info = []
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        if ws_f.sheet_state != "visible":
            continue
        sheets_info.append(analyze_sheet(ws_f, ws_v, resolver))
        print(f"  시트 분석: {name}")
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    out_md = os.path.join(out_dir, base + ".analysis.md")
    render_report(xlsx_path, sheets_info, out_md)
    print(f"보고서: {out_md}")
    return out_md


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python analyze_reference.py <실무.xlsx> [출력폴더]")
        sys.exit(2)
    target = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(target), "_analysis")
    analyze(target, out)
