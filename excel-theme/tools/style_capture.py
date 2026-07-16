# -*- coding: utf-8 -*-
"""
셀 스타일 추출 코어 — 샘플 피드백 diff(style_diff.py)와 참고파일 분석기
(analyze_reference.py)가 공유한다.

핵심:
  capture_cell / capture_region : 셀의 정규화된 스타일 dict
  build_theme_resolver          : 테마색(theme=N, tint=...) → 6자리 hex 환산
  normalize 계열                : Excel 저장 시 생기는 표현 차이(노이즈) 억제
    - 'FF000000' → '000000', 빈 테두리 → None, 기본 폰트색 None ≡ '000000' 등
    - 무편집 왕복(생성 → Excel에서 저장 → diff) 시 diff 0건이 목표.
"""
import colorsys
import re
import xml.etree.ElementTree as ET

from openpyxl.styles.colors import COLOR_INDEX

# 비교 대상 속성 (diff 보고서에 쓰는 한글 라벨)
ATTR_LABELS = {
    "fill_rgb":      "채우기색",
    "font_name":     "폰트",
    "font_size":     "글자 크기",
    "font_bold":     "굵게",
    "font_color":    "글자색",
    "border_top":    "위 테두리",
    "border_bottom": "아래 테두리",
    "border_left":   "왼쪽 테두리",
    "border_right":  "오른쪽 테두리",
    "number_format": "숫자서식",
    "align_h":       "가로 정렬",
    "align_v":       "세로 정렬",
    "wrap":          "줄바꿈",
}

# theme1.xml clrScheme 이름 → xlsx color theme 인덱스 순서
# (Excel 은 dk1/lt1, dk2/lt2 를 서로 바꿔 인덱싱한다)
_THEME_ORDER = ["lt1", "dk1", "lt2", "dk2",
                "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
                "hlink", "folHlink"]
_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def build_theme_resolver(wb):
    """워크북의 테마 XML을 파싱해 (theme_idx, tint) → 6자리 hex 함수를 만든다.
    테마 XML이 없으면 Office 기본 테마 색을 쓴다."""
    base = list(_OFFICE_DEFAULT)
    xml_src = getattr(wb, "loaded_theme", None)
    if xml_src:
        try:
            if isinstance(xml_src, bytes):
                xml_src = xml_src.decode("utf-8")
            root = ET.fromstring(xml_src)
            scheme = root.find(f".//{_A_NS}clrScheme")
            if scheme is not None:
                found = {}
                for child in scheme:
                    name = child.tag.replace(_A_NS, "")
                    srgb = child.find(f"{_A_NS}srgbClr")
                    sysc = child.find(f"{_A_NS}sysClr")
                    if srgb is not None:
                        found[name] = srgb.get("val", "000000").upper()
                    elif sysc is not None:
                        found[name] = sysc.get("lastClr", "000000").upper()
                for i, key in enumerate(_THEME_ORDER):
                    if key in found:
                        base[i] = found[key]
        except Exception:
            pass  # 파싱 실패 시 기본 테마로

    def resolve(theme_idx, tint):
        try:
            hex6 = base[int(theme_idx)]
        except (IndexError, ValueError, TypeError):
            return None
        return _apply_tint(hex6, tint or 0.0)

    return resolve


# Office 기본 테마(theme1.xml 부재 시 폴백) — lt1, dk1, lt2, dk2, accent1~6, hlink, folHlink
_OFFICE_DEFAULT = ["FFFFFF", "000000", "E7E6E6", "44546A",
                   "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47",
                   "0563C1", "954F72"]


def _apply_tint(hex6, tint):
    """MS tint 공식(HLS 명도 보정) 근사 적용."""
    if not tint:
        return hex6.upper()
    r = int(hex6[0:2], 16) / 255.0
    g = int(hex6[2:4], 16) / 255.0
    b = int(hex6[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1.0 + tint) if tint < 0 else l * (1.0 - tint) + tint
    r, g, b = colorsys.hls_to_rgb(h, min(max(l, 0.0), 1.0), s)
    return "%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


def resolve_color(color, resolver):
    """openpyxl Color → 6자리 hex 또는 None(자동/미지정)."""
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = color.rgb
        if not isinstance(rgb, str):
            return None
        return rgb[-6:].upper()
    if ctype == "theme":
        return resolver(color.theme, getattr(color, "tint", 0.0))
    if ctype == "indexed":
        idx = color.indexed
        if idx in (64, 65):       # 시스템 전경/배경 = 자동
            return None
        try:
            argb = COLOR_INDEX[idx]
            return argb[-6:].upper()
        except (IndexError, TypeError):
            return None
    return None  # auto 등


def _norm_font_color(hex6):
    """기본 글자색(None)과 명시적 검정은 시각적으로 동일 → '000000' 으로 통일."""
    return hex6 if hex6 else "000000"


# Excel 은 저장 시 숫자서식의 리터럴 문자를 백슬래시로 이스케이프해 다시 쓴다
#   '0.0x' → '0.0\x',  '(#,##0)' → '\(#,##0\)'  — 의미는 동일하므로 비교 전 제거.
def _norm_nf(fmt):
    if not fmt:
        return "General"
    return re.sub(r"\\(.)", r"\1", fmt)


# 테마 기본 폰트(scheme=minor/major)는 Excel 이 저장 시 이름을 테마 폰트로
# 바꿔 쓴다(Calibri → 맑은 고딕). 이름 대신 표지로 통일해 노이즈를 막는다.
_SCHEME_FONT = "(테마기본)"


def _norm_border_side(side, resolver):
    """Side → "style|hex" 문자열 또는 None. 색 미지정은 검정 취급."""
    if side is None or side.style is None:
        return None
    c = resolve_color(side.color, resolver) or "000000"
    return f"{side.style}|{c}"


def capture_cell(cell, resolver):
    """셀 1개의 정규화된 스타일 dict (ATTR_LABELS 의 키들)."""
    out = {}
    f = cell.fill
    if f is not None and f.patternType == "solid":
        out["fill_rgb"] = resolve_color(f.fgColor, resolver)
    else:
        out["fill_rgb"] = None

    fo = cell.font
    if getattr(fo, "scheme", None) in ("minor", "major"):
        out["font_name"] = _SCHEME_FONT
    else:
        out["font_name"] = fo.name or None
    out["font_size"] = float(fo.size) if fo.size else None
    out["font_bold"] = bool(fo.bold)
    out["font_color"] = _norm_font_color(resolve_color(fo.color, resolver))

    b = cell.border
    out["border_top"] = _norm_border_side(b.top, resolver)
    out["border_bottom"] = _norm_border_side(b.bottom, resolver)
    out["border_left"] = _norm_border_side(b.left, resolver)
    out["border_right"] = _norm_border_side(b.right, resolver)

    out["number_format"] = _norm_nf(cell.number_format)

    al = cell.alignment
    out["align_h"] = al.horizontal or "general"
    out["align_v"] = al.vertical or "bottom"
    out["wrap"] = bool(al.wrap_text)
    return out


def capture_coords(ws, coords, resolver):
    """coords 리스트의 셀들을 {coord: style_dict} 로."""
    return {coord: capture_cell(ws[coord], resolver) for coord in coords}


def capture_dims(ws, cols=(), rows=()):
    """열너비/행높이. 미지정(기본값)은 None.
    Excel 은 같은 너비의 인접 열을 <col min max> 범위 하나로 묶어 저장하므로
    범위를 펼쳐서 열별 너비 맵을 만든 뒤 조회한다."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    expanded = {}
    for d in ws.column_dimensions.values():
        if not d.width:
            continue
        lo = d.min or column_index_from_string(d.index)
        hi = d.max or lo
        for ci in range(lo, hi + 1):
            expanded[get_column_letter(ci)] = round(d.width, 2)
    cw = {letter: expanded.get(letter) for letter in cols}
    rh = {}
    for r in rows:
        d = ws.row_dimensions.get(r)
        rh[str(r)] = round(d.height, 2) if (d is not None and d.height) else None
    return {"col_widths": cw, "row_heights": rh}


def diff_styles(old, new, attrs=None):
    """두 스타일 dict 비교 → {attr: (old, new)} (다른 것만)."""
    out = {}
    for attr in (attrs or ATTR_LABELS):
        ov, nv = old.get(attr), new.get(attr)
        if attr == "font_size" and ov is not None and nv is not None:
            if abs(float(ov) - float(nv)) < 0.01:
                continue
        if ov != nv:
            out[attr] = (ov, nv)
    return out
