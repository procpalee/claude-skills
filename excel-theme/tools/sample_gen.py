# -*- coding: utf-8 -*-
"""
피드백용 샘플 워크북 + 스냅샷 생성기.

    python sample_gen.py <out_dir>
  → <out_dir>\samples_<theme>.xlsx          (테마당 1개, 시트 12종)
  → <out_dir>\samples_<theme>.snapshot.json (셀 역할맵 + 생성 시점 스타일)

시트 = examples/table_samples.py 의 SAMPLES 8종 + tools/sample_specs.py 4종.
생성기가 각 셀의 '역할'(title/header/body/section/subtotal/total/input/...)을
직접 알고 기록하므로, style_diff.py 가 추론 없이 규칙 단위로 비교할 수 있다.

사용 흐름: 생성 → 사용자가 엑셀에서 직접 서식 수정·저장 → dev.ps1 -Diff
"""
import json
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)                       # excel-theme/
sys.path.insert(0, _ROOT)
sys.path.insert(0, os.path.join(_ROOT, "examples"))
sys.path.insert(0, _HERE)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from excel_theme import (THEMES, NUMBER_FORMATS, apply_theme, get_theme,
                         mark_cells, style_subheader_row, style_total_row)
from table_samples import SAMPLES
from sample_specs import EXTRA_SAMPLES
from style_capture import build_theme_resolver, capture_coords, capture_dims

SNAPSHOT_VERSION = 1


def _expand_range(rng):
    """'C5:D8' / 'C5' → 좌표 리스트."""
    c1, r1, c2, r2 = range_boundaries(rng)
    return [f"{get_column_letter(c)}{r}"
            for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]


def render_table(ws, spec, theme_name, *, top_row, with_title):
    """표 1개를 그리고 (roles, fmt_keys, last_row, max_col) 반환.
    examples/build_preview.py 의 build_sheet 와 같은 규칙 + 역할 기록."""
    th = get_theme(theme_name)
    roles, fmt_keys = {}, {}
    headers = spec["headers"]
    n_cols = len(headers)
    max_col = 1 + n_cols                       # B(2)부터 → 마지막 열 인덱스
    last_col_letter = get_column_letter(max_col)

    group_header = spec.get("group_header")
    if group_header:
        col = 2
        for label, span in group_header:
            ws.cell(row=top_row, column=col, value=label)
            if span > 1:
                ws.merge_cells(start_row=top_row, start_column=col,
                               end_row=top_row, end_column=col + span - 1)
            col += span
        header_row = top_row + 1
        range_top = top_row
        for c in range(2, max_col + 1):
            roles[f"{get_column_letter(c)}{top_row}"] = "group_header"
    else:
        header_row = top_row
        range_top = top_row

    for j, h in enumerate(headers, start=2):
        ws.cell(row=header_row, column=j, value=h)
        roles[f"{get_column_letter(j)}{header_row}"] = "header"

    data_start = header_row + 1
    rows = spec["rows"]
    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=2):
            if val is not None:
                ws.cell(row=data_start + i, column=j, value=val)
    last_row = data_start + len(rows) - 1

    zebra = spec.get("zebra", th.get("excel_zebra", True))
    hl_neg = spec.get("highlight_neg", th.get("excel_highlight_neg", True))
    apply_theme(
        ws, theme=theme_name, header_row=header_row,
        data_range=f"B{range_top}:{last_col_letter}{last_row}",
        title_cell=("B2" if with_title else None),
        number_format_cols=spec.get("number_format_cols"),
        zebra=zebra, freeze=spec.get("freeze", False), highlight_neg=hl_neg,
    )

    if with_title:
        roles["B2"] = "title"
        if th.get("title_fill"):
            for c in range(3, max_col + 1):
                roles[f"{get_column_letter(c)}2"] = "title_bar"

    if group_header:
        style_subheader_row(ws, th, row=top_row, min_col=2, max_col=max_col)

    # 본문 역할 (zebra 줄은 body_zebra)
    for i, r in enumerate(range(data_start, last_row + 1)):
        role = "body_zebra" if (zebra and i % 2 == 1) else "body"
        for c in range(2, max_col + 1):
            roles[f"{get_column_letter(c)}{r}"] = role

    def sheet_row(idx):
        return data_start + idx - 1

    for idx in spec.get("subheader_rows", []):
        style_subheader_row(ws, th, row=sheet_row(idx), min_col=2, max_col=max_col)
        for c in range(2, max_col + 1):
            roles[f"{get_column_letter(c)}{sheet_row(idx)}"] = "section"
    for idx in spec.get("subtotal_rows", []):
        style_total_row(ws, th, row=sheet_row(idx), min_col=2, max_col=max_col, fill=False)
        for c in range(2, max_col + 1):
            roles[f"{get_column_letter(c)}{sheet_row(idx)}"] = "subtotal"
    for idx in spec.get("total_rows", []):
        style_total_row(ws, th, row=sheet_row(idx), min_col=2, max_col=max_col, fill=True)
        for c in range(2, max_col + 1):
            roles[f"{get_column_letter(c)}{sheet_row(idx)}"] = "total"

    # 입력/연결 셀 (역할 덮어쓰기 — 채우기 비교용)
    inputs = spec.get("input_cells", [])
    linked = spec.get("linked_cells", [])
    if inputs or linked:
        mark_cells(ws, th, input=list(inputs), linked=list(linked))
        for rng in inputs:
            for coord in _expand_range(rng):
                roles[coord] = "input"
        for rng in linked:
            for coord in _expand_range(rng):
                roles[coord] = "linked"

    # 숫자서식 키 기록 (NUMBER_FORMATS 의 키인 것만 — 서식 변경 → 키 갱신 제안용)
    for col_letter, key in (spec.get("number_format_cols") or {}).items():
        if key in NUMBER_FORMATS:
            for r in range(data_start, last_row + 1):
                fmt_keys[f"{col_letter}{r}"] = key

    return roles, fmt_keys, last_row, max_col


def render_sheet(ws, spec, theme_name):
    """시트 1개(표 1~N개) 렌더 + 역할/서식키/치수 메타 반환."""
    roles, fmt_keys = {}, {}
    specs = spec.get("tables") or [spec]
    top_row, overall_max_col, last_row = 4, 2, 4
    for i, tspec in enumerate(specs):
        if i == 0 and spec.get("title"):
            ws["B2"] = spec["title"]
        r, f, last_row, mc = render_table(
            ws, tspec, theme_name, top_row=top_row,
            with_title=(i == 0 and bool(spec.get("title"))))
        roles.update(r)
        fmt_keys.update(f)
        overall_max_col = max(overall_max_col, mc)
        top_row = last_row + 3                  # 표 간 간격 2행
    dims_cols = [get_column_letter(c) for c in range(1, overall_max_col + 1)]
    dims_rows = list(range(1, last_row + 1))
    return roles, fmt_keys, dims_cols, dims_rows


def generate(theme_name, out_dir):
    """테마 1개 → (xlsx 경로, snapshot 경로)."""
    wb = Workbook()
    wb.remove(wb.active)
    meta = {}
    for spec in SAMPLES + EXTRA_SAMPLES:
        ws = wb.create_sheet(spec["name"][:31])
        roles, fmt_keys, dims_cols, dims_rows = render_sheet(ws, spec, theme_name)
        meta[ws.title] = {"roles": roles, "fmt_keys": fmt_keys,
                          "dims_cols": dims_cols, "dims_rows": dims_rows}

    xlsx_path = os.path.join(out_dir, f"samples_{theme_name}.xlsx")
    snap_path = os.path.join(out_dir, f"samples_{theme_name}.snapshot.json")
    wb.save(xlsx_path)

    # 저장본을 다시 읽어 스냅샷 추출 → '파일에 실제로 기록된 상태'와 일치 보장
    rb = load_workbook(xlsx_path)
    resolver = build_theme_resolver(rb)
    sheets = {}
    for name, m in meta.items():
        ws = rb[name]
        sheets[name] = {
            "roles": m["roles"],
            "fmt_keys": m["fmt_keys"],
            "styles": capture_coords(ws, list(m["roles"]), resolver),
            "dims": capture_dims(ws, cols=m["dims_cols"], rows=m["dims_rows"]),
        }
    snapshot = {"version": SNAPSHOT_VERSION, "theme": theme_name,
                "file": os.path.basename(xlsx_path), "sheets": sheets}
    with open(snap_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=1)
    return xlsx_path, snap_path


def main(out_dir):
    os.makedirs(out_dir, exist_ok=True)
    print("피드백용 샘플 생성 중...")
    for theme_name in THEMES:
        xlsx, snap = generate(theme_name, out_dir)
        print(f"  생성: {xlsx}")
        print(f"        {os.path.basename(snap)}")
    print("완료. 엑셀에서 서식을 직접 고쳐 저장한 뒤  dev.bat -Diff  를 실행하세요.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python sample_gen.py <출력폴더>")
        sys.exit(2)
    main(sys.argv[1])
