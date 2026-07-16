# -*- coding: utf-8 -*-
"""
엑셀 테마 — 색/폰트/표 규칙 + openpyxl 헬퍼를 **이 한 파일**에 모았다.
(office-palette 분리 없이 단독 완결. 워드/PPT 스킬도 이 파일의 팔레트를 참조한다.)

■ 색/폰트/표 규칙을 바꾸려면 아래 "편집 영역"의 PALETTES / NUMBER_FORMATS / DEFAULT_THEME 만 고친다. ■
그 아래 매핑·헬퍼 로직은 보통 손댈 필요가 없다.

    from excel_theme import apply_theme
    apply_theme(ws, theme="default", header_row=4, data_range="B4:F9",
                title_cell="B2", number_format_cols={"C": "accounting"})
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.utils.cell import coordinate_from_string

# ══════════════════════════════════════════════════════════════════════
# ■ 최신 Office 테마 강제 — openpyxl 기본 "Office 2007-2010" 제거 ■
# ══════════════════════════════════════════════════════════════════════
# openpyxl 은 저장 시 항상 구버전(Office 2007-2010) 테마 XML 을 끼워 넣어,
# Excel 의 [페이지 레이아웃]>[테마] 가 "Office 2007-2010" 으로 잡힌다.
# 아래에서 최신(2013+) Office 테마(Calibri Light/Calibri · 모던 색상)를 강제한다.
#   - 이 모듈을 import 하면 openpyxl 기본 저장 경로가 자동으로 최신 테마를 쓴다.
#   - 특정 워크북만 개별 지정하려면 modernize_theme(wb) 를 save() 직전에 호출.
MODERN_OFFICE_THEME = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Office Theme">'
    '<a:themeElements><a:clrScheme name="Office">'
    '<a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>'
    '<a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>'
    '<a:dk2><a:srgbClr val="44546A"/></a:dk2><a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>'
    '<a:accent1><a:srgbClr val="4472C4"/></a:accent1><a:accent2><a:srgbClr val="ED7D31"/></a:accent2>'
    '<a:accent3><a:srgbClr val="A5A5A5"/></a:accent3><a:accent4><a:srgbClr val="FFC000"/></a:accent4>'
    '<a:accent5><a:srgbClr val="5B9BD5"/></a:accent5><a:accent6><a:srgbClr val="70AD47"/></a:accent6>'
    '<a:hlink><a:srgbClr val="0563C1"/></a:hlink><a:folHlink><a:srgbClr val="954F72"/></a:folHlink>'
    '</a:clrScheme><a:fontScheme name="Office">'
    '<a:majorFont><a:latin typeface="맑은 고딕"/>'
    '<a:ea typeface="맑은 고딕"/><a:cs typeface=""/></a:majorFont>'
    '<a:minorFont><a:latin typeface="맑은 고딕"/>'
    '<a:ea typeface="맑은 고딕"/><a:cs typeface=""/></a:minorFont></a:fontScheme>'
    '<a:fmtScheme name="Office">'
    '<a:fillStyleLst><a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
    '<a:gradFill rotWithShape="1"><a:gsLst>'
    '<a:gs pos="0"><a:schemeClr val="phClr"><a:lumMod val="110000"/><a:satMod val="105000"/><a:tint val="67000"/></a:schemeClr></a:gs>'
    '<a:gs pos="50000"><a:schemeClr val="phClr"><a:lumMod val="105000"/><a:satMod val="103000"/><a:tint val="73000"/></a:schemeClr></a:gs>'
    '<a:gs pos="100000"><a:schemeClr val="phClr"><a:lumMod val="105000"/><a:satMod val="109000"/><a:tint val="81000"/></a:schemeClr></a:gs>'
    '</a:gsLst><a:lin ang="5400000" scaled="0"/></a:gradFill>'
    '<a:gradFill rotWithShape="1"><a:gsLst>'
    '<a:gs pos="0"><a:schemeClr val="phClr"><a:satMod val="103000"/><a:lumMod val="102000"/><a:tint val="94000"/></a:schemeClr></a:gs>'
    '<a:gs pos="50000"><a:schemeClr val="phClr"><a:satMod val="110000"/><a:lumMod val="100000"/><a:shade val="100000"/></a:schemeClr></a:gs>'
    '<a:gs pos="100000"><a:schemeClr val="phClr"><a:lumMod val="99000"/><a:satMod val="120000"/><a:shade val="78000"/></a:schemeClr></a:gs>'
    '</a:gsLst><a:lin ang="5400000" scaled="0"/></a:gradFill></a:fillStyleLst>'
    '<a:lnStyleLst>'
    '<a:ln w="6350" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/><a:miter lim="800000"/></a:ln>'
    '<a:ln w="12700" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/><a:miter lim="800000"/></a:ln>'
    '<a:ln w="19050" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/><a:miter lim="800000"/></a:ln>'
    '</a:lnStyleLst>'
    '<a:effectStyleLst><a:effectStyle><a:effectLst/></a:effectStyle><a:effectStyle><a:effectLst/></a:effectStyle>'
    '<a:effectStyle><a:effectLst><a:outerShdw blurRad="57150" dist="19050" dir="5400000" rotWithShape="0">'
    '<a:srgbClr val="000000"><a:alpha val="63000"/></a:srgbClr></a:outerShdw></a:effectLst></a:effectStyle></a:effectStyleLst>'
    '<a:bgFillStyleLst><a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
    '<a:solidFill><a:schemeClr val="phClr"><a:tint val="95000"/><a:satMod val="170000"/></a:schemeClr></a:solidFill>'
    '<a:gradFill rotWithShape="1"><a:gsLst>'
    '<a:gs pos="0"><a:schemeClr val="phClr"><a:tint val="93000"/><a:satMod val="150000"/><a:shade val="98000"/><a:lumMod val="102000"/></a:schemeClr></a:gs>'
    '<a:gs pos="50000"><a:schemeClr val="phClr"><a:tint val="98000"/><a:satMod val="130000"/><a:shade val="90000"/><a:lumMod val="103000"/></a:schemeClr></a:gs>'
    '<a:gs pos="100000"><a:schemeClr val="phClr"><a:shade val="63000"/><a:satMod val="120000"/></a:schemeClr></a:gs>'
    '</a:gsLst><a:lin ang="5400000" scaled="0"/></a:gradFill></a:bgFillStyleLst></a:fmtScheme>'
    '</a:themeElements><a:objectDefaults/><a:extraClrSchemeLst/></a:theme>'
)


def modernize_theme(wb):
    """워크북에 최신 Office 테마를 박아 넣는다(save 직전 호출). 반환: wb."""
    wb.loaded_theme = MODERN_OFFICE_THEME
    return wb


# import 시점에 openpyxl 기본 저장 경로(loaded_theme 미설정 시)를 최신 테마로 교체.
# writer.excel 이 `from .theme import theme_xml` 로 값을 복사해 가므로 그 네임스페이스를 직접 패치.
try:
    import openpyxl.writer.excel as _ox_writer
    _ox_writer.theme_xml = MODERN_OFFICE_THEME
except Exception:
    pass

# ══════════════════════════════════════════════════════════════════════
# ■■■ 편집 영역 — 색 / 폰트 / 표 규칙은 여기서 모두 수정한다 ■■■
# ══════════════════════════════════════════════════════════════════════
#
# 의미 토큰(색): primary 헤더/제목·테두리 / primary_dark 주요항목·합계 글자 /
#   secondary·secondary_lt 보조 / band 합계·강조 배경 / subheader 섹션배경 /
#   surface 옅은 면 / text 본문글자 / note 강조박스 / accent 포인트 /
#   negative 음수 / header_font 헤더글자 / title_color 제목글자 / border_in·border_out 테두리
# 엑셀 폰트는 font_excel(없으면 font_name). 엑셀 렌더 플래그:
#   excel_table_style: "rules"(흰배경 헤더+가로구분선) | "grid"(전체 격자)
#   excel_zebra / excel_highlight_neg / excel_neg_red (음수 [Red] 사용)

# 미지정 시 기본 테마. 공식 카탈로그 3종 — 목적에 따라 theme= 로 골라 쓴다:
#   default = 정산표·결산(회색·기본) / procpa = 브랜드 블루(procpa.co.kr) / dcf-valuation = DCF·평가
#   navy·charcoal 은 legacy(카탈로그 제외, 하위호환용 — 기존 감사조서 룩 유지).
#   구 이름은 ALIASES 로 계속 동작: closing→default, valuation→dcf-valuation,
#   audit_charcoal→charcoal, audit_navy→navy (감사조서 빌더 frame.py 하위호환)
DEFAULT_THEME = "default"

PALETTES = {
    # ── DCF·평가용 : Ocean Blue Clean (구 valuation) ─────────────────────
    "dcf-valuation": {
        "label": "DCF·평가 (Ocean Blue Clean)",
        "font_name": "Pretendard",       # 워드/PPT 기본 폰트
        "font_excel": "맑은 고딕",         # 엑셀 폰트
        "font_fallback": "맑은 고딕",
        "sizes": {"title": 20, "subtitle": 13, "h1": 16, "h2": 13,
                  "h3": 11, "header": 11, "body": 11, "small": 9},
        "excel_zebra": False, "excel_highlight_neg": False, "excel_neg_red": False,
        "excel_table_style": "rules",
        "excel_title_size": 12, "excel_size_header": 11, "excel_size_body": 11,
        "excel_row_header": 22, "excel_row_body": 18,
        "colors": {
            "primary": "004889", "primary_dark": "1F3864", "secondary": "2B579A",
            "secondary_lt": "5B9BD5", "band": "F5F8FC", "subheader": "F5F8FC",
            "surface": "F5F8FC", "text": "3D4F5F", "note": "F5F8FC",
            "accent": "5B9BD5", "negative": "5A6A7A", "header_font": "FFFFFF",
            "title_color": "1F3864", "border_in": "D6E4F0", "border_out": "004889",
        },
    },
    # ── 정산표·결산·감사조서 : 삼보모터스 정산표 (구 settlement) ────────
    #   statement 스타일: 회색(F2F2F2) bold 헤더 + 표 상/하 medium 테두리 + 격자 없음.
    #   하드코딩 셀 #FFFBEF, 음수 빨강 괄호, 맑은 고딕. (FY26.1Q 파일 분석)
    "default": {
        "label": "default (정산표·결산 — 회색 statement)",
        "font_name": "맑은 고딕", "font_excel": "맑은 고딕", "font_fallback": "Malgun Gothic",
        "sizes": {"title": 12, "subtitle": 11, "h1": 14, "h2": 12,
                  "h3": 11, "header": 11, "body": 11, "small": 10},
        "excel_zebra": False, "excel_highlight_neg": False, "excel_neg_red": True,
        "excel_table_style": "frame", "excel_title_fill": True,
        "excel_currency_format": "million_won",   # 통화 기본 단위 = 백만원 (정산표 관행)
        "excel_title_size": 12, "excel_size_header": 11, "excel_size_body": 11,
        "excel_row_header": 20, "excel_row_body": 16,
        "colors": {
            "primary": "404040",       # (워드/PPT 헤딩·헤더·타이틀) 진회색
            "primary_dark": "262626",  # 강조/소계·합계 글자
            "secondary": "808080", "secondary_lt": "BFBFBF",
            "band": "F2F2F2",          # 줄무늬 배경(현재 미사용)
            "subheader": "FFFFFF",     # 섹션행 배경 없음(굵은 글씨로만 구분)
            "surface": "FFFFFF",
            "text": "000000",          # 본문 검정
            "note": "FFFBEF",          # 하드코딩(입력) 셀 — 크림
            "linked": "F2F2F2",        # 참조값(타시트 연결) 셀 — 연회색
            "accent": "C00000",        # 포인트(진한 빨강)
            "negative": "FF0000",      # 음수 빨강
            "header_font": "FFFFFF",   # (워드/PPT 헤더 글자) 흰색
            "title_color": "000000",
            "border_in": "BFBFBF",     # (쓰면) 얇은 선
            "border_out": "000000",    # 엑셀 표 상/하 굵은(medium) 선
            # ── 엑셀 전용 오버라이드 (정산표 statement 룩) ──
            "excel_header_fill": "F2F2F2",  # 본문 기본 헤더(회색)
            "excel_header_text": "000000",  # 헤더 글자(검정 bold)
            "excel_header2_fill": "DDE3E8",  # 본문 세컨더리 헤더(연청회색)
            "excel_header2_text": "000000",
            "excel_total_fill": "F2F2F2",    # 소계/합계 배경 = 기본 헤더색
            "title_bg": "393939",           # 제목(B2) 헤더 바(진회색, 데이터 끝열까지)
            "title_font_color": "FFFFFF",
        },
    },

    # ── PROCPA 브랜드 : default 의 statement 레이아웃 + procpa.co.kr 블루 팔레트 ──
    #   딥네이비(#0B1C4A) 타이틀 바 + 블루 틴트(#EFF5FF) 헤더 + 프라이머리 블루(#2563EB) 악센트.
    #   입력(note)·연결(linked)·음수(negative)는 기능색이라 default 와 동일하게 유지.
    "procpa": {
        "label": "procpa (브랜드 블루 — 딥네이비 & 블루)",
        "font_name": "Pretendard", "font_excel": "맑은 고딕", "font_fallback": "맑은 고딕",
        "sizes": {"title": 12, "subtitle": 11, "h1": 14, "h2": 12,
                  "h3": 11, "header": 11, "body": 11, "small": 10},
        "excel_zebra": False, "excel_highlight_neg": False, "excel_neg_red": True,
        "excel_table_style": "frame", "excel_title_fill": True,
        "excel_currency_format": "million_won",
        "excel_title_size": 12, "excel_size_header": 11, "excel_size_body": 11,
        "excel_row_header": 20, "excel_row_body": 16,
        "colors": {
            "primary": "2563EB",       # (워드/PPT 헤딩·헤더) 프라이머리 블루
            "primary_dark": "0B1C4A",  # 강조/소계·합계 글자 — 딥네이비
            "secondary": "4A5160", "secondary_lt": "94A3B8",
            "band": "EFF5FF",          # 줄무늬/합계 배경 — 블루 틴트
            "subheader": "FFFFFF",     # 섹션행 배경 없음(굵은 글씨로만 구분)
            "surface": "FFFFFF",
            "text": "000000",
            "note": "FFFBEF",          # 하드코딩(입력) 셀 — 크림 (기능색, 전 테마 공통)
            "linked": "F2F2F2",        # 참조값(타시트 연결) 셀 — 연회색 (기능색)
            "accent": "2563EB",        # 포인트
            "negative": "FF0000",      # 음수 빨강 (회계 관행)
            "header_font": "FFFFFF",
            "title_color": "0B1C4A",
            "border_in": "C9D4E5",     # 블루그레이 헤어라인
            "border_out": "0B1C4A",    # 표 상/하 medium 선 — 딥네이비
            # ── 엑셀 전용 오버라이드 (statement 룩) ──
            "excel_header_fill": "EFF5FF",   # 본문 기본 헤더(블루 틴트)
            "excel_header_text": "0B1C4A",
            "excel_header2_fill": "F3F5F8",  # 본문 세컨더리 헤더(그레이 서피스)
            "excel_header2_text": "0B1C4A",
            "excel_total_fill": "EFF5FF",    # 소계/합계 배경 = 기본 헤더색
            "title_bg": "0B1C4A",            # 제목(B2) 헤더 바(딥네이비)
            "title_font_color": "FFFFFF",
        },
    },

    # ── [legacy] 감사조서용 팔레트 2종 — 카탈로그 제외, 하위호환용(기존 조서 룩 유지) ──
    #   공통: 맑은 고딕, 입력=크림(note)·연결=옅은 동색(linked), grid 테두리(외곽 medium=진한색).
    #   고도화(2026-06): 섹션·헤더 계층 대비 정교화 + 테마별 linked 색 추가.
    # 표=frame(삼선표: 좌우 외곽선 없음·상하 굵게·내부 얇게) · 헤더 3색(제목/본문/세컨더리)
    # 셀배경: 하드코딩=note(크림) / 수식=흰배경(없음) / 참조값=linked(#F2F2F2)
    "navy": {
        "label": "navy (네이비 & 골드) — legacy",
        "legacy": True,
        "font_name": "맑은 고딕", "font_excel": "맑은 고딕", "font_fallback": "Malgun Gothic",
        "sizes": {"title": 12, "subtitle": 11, "h1": 14, "h2": 12, "h3": 11, "header": 11, "body": 11, "small": 9},
        # excel_highlight_neg=False → 음수는 숫자서식 [Red](기본 빨강)만, bold/별색 오버레이 없음
        "excel_zebra": False, "excel_highlight_neg": False, "excel_neg_red": True,
        "excel_table_style": "frame", "excel_title_fill": True,
        "excel_title_size": 12, "excel_size_header": 11, "excel_size_body": 11,
        "excel_row_header": 22, "excel_row_body": 20,
        "colors": {
            "primary": "1F3A5F", "primary_dark": "16293F", "secondary": "3A5172", "secondary_lt": "8AA0BC",
            "band": "EEF2F7", "subheader": "FFFFFF", "surface": "FFFFFF", "text": "1A2330",
            "note": "FFFBEF", "linked": "F2F2F2", "accent": "C9A227", "negative": "FF0000", "header_font": "FFFFFF",
            "title_color": "1F3A5F", "border_in": "C5CEDA", "border_out": "1F3A5F",
            "excel_header_fill": "E1E8F1", "excel_header_text": "16293F",      # 본문 기본 헤더
            "excel_header2_fill": "DCE6F4", "excel_header2_text": "16293F",    # 본문 세컨더리 헤더
            "excel_total_fill": "FFFFFF",                                      # 소계/합계 배경 없음(윗선+굵게로만)
            "title_bg": "1F3A5F", "title_font_color": "FFFFFF",               # 제목(B2) 헤더
            # subheader=FFFFFF → 섹션행 별도 배경색 없음(굵은 글씨로만 구분)
        },
    },
    "charcoal": {
        "label": "charcoal (차콜 & 틸) — legacy",
        "legacy": True,
        "font_name": "맑은 고딕", "font_excel": "맑은 고딕", "font_fallback": "Malgun Gothic",
        "sizes": {"title": 12, "subtitle": 11, "h1": 14, "h2": 12, "h3": 11, "header": 11, "body": 11, "small": 9},
        # excel_highlight_neg=False → 음수는 굵게/별색 오버레이 없이 숫자서식 [Red](기본 빨강)만
        "excel_zebra": False, "excel_highlight_neg": False, "excel_neg_red": True,
        "excel_table_style": "frame", "excel_title_fill": True,
        "excel_title_size": 12, "excel_size_header": 11, "excel_size_body": 11,
        "excel_row_header": 22, "excel_row_body": 20,
        "colors": {
            "primary": "2A2E35", "primary_dark": "1A1D22", "secondary": "4A4F58", "secondary_lt": "9AA0AA",
            "band": "EDEFF2", "subheader": "FFFFFF", "surface": "FFFFFF", "text": "22262B",
            "note": "FFFBEF", "linked": "F2F2F2", "accent": "2E8B7F", "negative": "FF0000", "header_font": "FFFFFF",
            "title_color": "2A2E35", "border_in": "D2D6DC", "border_out": "2A2E35",
            "excel_header_fill": "D6DEE6", "excel_header_text": "2A2E35",      # 본문 기본 헤더(슬레이트)
            "excel_header2_fill": "E8ECEF", "excel_header2_text": "2A2E35",    # 본문 세컨더리(연그레이)
            "excel_total_fill": "FFFFFF",                                      # 소계/합계 배경 없음(윗선+굵게로만)
            "title_bg": "2A2E35", "title_font_color": "FFFFFF",               # 제목(B2) 헤더
            # subheader=FFFFFF → 섹션행 별도 배경색 없음(굵은 글씨로만 구분)
        },
    },
}

# 구 이름 → 신 이름 별칭 (감사조서 빌더 frame.py·valuation-tools 등 하위호환)
ALIASES = {"valuation": "dcf-valuation",
           "audit_charcoal": "charcoal", "audit_navy": "navy", "closing": "default",
           "default - charcoal": "charcoal", "default - navy": "navy"}

# 공통 숫자 서식 (음수=빨강 괄호, 0=대시). 테마가 [Red] 제거 옵션을 가질 수 있음.
NUMBER_FORMATS = {
    "accounting":   "#,###,##0;[Red](#,###,##0);-",
    "percent_acct": '0.00%_);[Red](0.00%);"-"_);@_)',
    "thousands":    "#,##0;[Red](#,##0)",
    "percent":      "0.0%;[Red]-0.0%",
    "currency_won": '#,##0"원";[Red](#,##0"원")',
    "million_won":  '#,##0,,"백만원";[Red](#,##0,,"백만원")',
    "decimal":      "#,##0.00;[Red](#,##0.00)",
    "date":         "yyyy-mm-dd",
    # 추가: 모델링/배수/스케일/증감
    "multiple":     "0.0x",                                  # 배수 (8.5x)
    "million":      "#,##0,,;[Red](#,##0,,);-",              # 백만 스케일(라벨 없음)
    "billion":      "#,##0,,,;[Red](#,##0,,,);-",            # 십억 스케일
    "change":       "+#,##0;[Red]-#,##0;-",                  # 증감(부호)
    "bp":           '#,##0" bp"',                            # 베이시스포인트
}
# ══════════════════════════════════════════════════════════════════════
# ■■■ 편집 영역 끝 — 아래는 로직 (보통 수정 불필요) ■■■
# ══════════════════════════════════════════════════════════════════════

__all__ = [
    "apply_theme", "style_header_row", "apply_borders",
    "set_number_formats", "zebra_stripes", "freeze_below_header",
    "autofit_columns", "highlight_negatives", "highlight_threshold",
    "style_total_row", "style_subheader_row", "highlight_hardcoded",
    "write_table", "mark_cells",
    "PALETTES", "NUMBER_FORMATS", "THEMES", "DEFAULT_THEME",
    "get_palette", "get_theme", "color",
    "MODERN_OFFICE_THEME", "modernize_theme",
]


# ── 팔레트 접근 (워드/PPT 도 사용: 의미 토큰 dict 그대로) ─────────────
def _canon(name):
    """구 이름(ALIASES)을 신 이름으로 정규화."""
    return ALIASES.get(name, name)


def get_palette(name=None):
    """팔레트(의미 토큰) dict 반환. None/미존재 시 기본. (구 이름 별칭 허용)"""
    if not name:
        return PALETTES[DEFAULT_THEME]
    name = _canon(name)
    if name not in PALETTES:
        raise KeyError(f"알 수 없는 테마 '{name}'. 사용 가능: {', '.join(PALETTES)}")
    return PALETTES[name]


def color(name, role):
    """팔레트 name 의 의미색 role 을 hex 로."""
    return get_palette(name)["colors"][role]


# ── 의미 팔레트 → 엑셀 토큰(header_bg 등) 매핑 ───────────────────────
def _excel_theme(palette_name):
    p = PALETTES[palette_name]
    c, s = p["colors"], p["sizes"]
    nf = dict(NUMBER_FORMATS)
    if not p.get("excel_neg_red", True):           # 음수 [Red] 제거(괄호만)
        nf = {k: v.replace("[Red]", "") for k, v in nf.items()}
    style = p.get("excel_table_style", "grid")
    rules_like = style in ("rules", "statement")     # 격자 없음·합계 가로선 방식
    return {
        "font_name":          p.get("font_excel", p["font_name"]),
        "font_size_title":    p.get("excel_title_size", 12),
        "font_size_header":   p.get("excel_size_header", s["header"]),
        "font_size_body":     p.get("excel_size_body", s["body"]),
        "header_bg":          c.get("excel_header_fill", c["primary"]),
        "header_font_color":  c.get("excel_header_text", c["header_font"]),
        "header2_bg":         c.get("excel_header2_fill", c["subheader"]),    # 본문 세컨더리 헤더
        "header2_font":       c.get("excel_header2_text", c["text"]),
        "title_color":        c["title_color"],
        "band_fill":          c["band"],
        "subheader_fill":     c["subheader"],
        "total_fill":         c.get("excel_total_fill", c["band"]),   # 소계/합계 배경(기본=band, 테마 지정 시 헤더색 등)
        "note_fill":          c["note"],
        "linked_fill":        c.get("linked", "CCECFF"),     # 타시트 연결 셀(파랑)
        "currency_format":    p.get("excel_currency_format", "accounting"),
        "accent_color":       c["negative"],
        "border_color":       c["border_in"],
        "border_color_outer": c["border_out"],
        "row_height_header":  p.get("excel_row_header", 22),
        "row_height_body":    p.get("excel_row_body", 18),
        "number_formats":     nf,
        "excel_zebra":         p.get("excel_zebra", True),
        "excel_highlight_neg": p.get("excel_highlight_neg", True),
        "header_fill":         style != "rules",
        "header_text_color":   (c["primary"] if style == "rules"
                                else c.get("excel_header_text", c["header_font"])),
        "border_mode":         style,
        "text_color":          c["text"],
        "title_fill":          p.get("excel_title_fill", style == "rules"),
        "title_bg":            c.get("title_bg", c["primary"]),
        "title_font_color":    c.get("title_font_color", c["header_font"]),
        "total_font_color":    (c["primary_dark"] if rules_like else None),
        "legacy":              p.get("legacy", False),   # 카탈로그 제외(하위호환) 표시
    }


THEMES = {name: _excel_theme(name) for name in PALETTES}


def get_theme(name=None):
    """엑셀 토큰 dict 반환. None/미존재 시 기본. (구 이름 별칭 허용)"""
    if not name:
        return THEMES[DEFAULT_THEME]
    name = _canon(name)
    if name not in THEMES:
        raise KeyError(f"알 수 없는 테마 '{name}'. 사용 가능: {', '.join(THEMES)}")
    return THEMES[name]


# ──────────────────────────────────────────────────────────────
# 내부 유틸
# ──────────────────────────────────────────────────────────────
def _bounds(ws, data_range):
    """data_range 문자열(또는 None)을 (min_col,min_row,max_col,max_row)로.
    None 이면 ws.dimensions(사용 영역)를 쓴다."""
    if data_range is None:
        data_range = ws.calculate_dimension()
    return range_boundaries(data_range)


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# ──────────────────────────────────────────────────────────────
# 1. 헤더 스타일
# ──────────────────────────────────────────────────────────────
def style_header_row(ws, theme, header_row, min_col, max_col, secondary=False):
    """헤더 행 스타일.

    header_fill=True  : 배경색 채움 + 흰(또는 지정) 글씨 (격자 스타일, 기본)
    header_fill=False : 흰 배경 + primary 굵은 글씨 + 하단 medium 선만
                        (Ocean Blue '가로 구분선' 스타일)
    secondary=True    : 본문 세컨더리 헤더 색(header2_bg/header2_font) — 한 시트에 표가
                        2차로 나올 때 본문 기본 헤더와 구분.
    """
    mode = theme.get("border_mode", "grid")
    header_fill = theme.get("header_fill", True)
    if secondary:
        bg = theme.get("header2_bg", theme["header_bg"])
        text_color = theme.get("header2_font",
                               theme.get("header_text_color", theme["header_font_color"]))
    else:
        bg = theme["header_bg"]
        text_color = theme.get("header_text_color", theme["header_font_color"])
    oc = theme["border_color_outer"]
    font = Font(name=theme["font_name"], size=theme["font_size_header"],
                bold=True, color=text_color)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill = _fill(bg) if header_fill else None
    # 헤더 테두리: rules(흰헤더)=하단 medium / statement=상단 medium + 하단 thin
    top_b = bot_b = None
    if mode == "rules" and not header_fill:
        bot_b = Side(style="medium", color=oc)
    elif mode == "statement":
        top_b = Side(style="medium", color=oc)
        bot_b = Side(style="thin", color=oc)
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=header_row, column=col)
        if fill is not None:
            c.fill = fill
        c.font = font
        c.alignment = align
        if top_b is not None or bot_b is not None:
            c.border = Border(top=top_b, bottom=bot_b)
    ws.row_dimensions[header_row].height = None   # 행높이 자동맞춤(헤더 wrap 시 자동 확장)


# ──────────────────────────────────────────────────────────────
# 2. 테두리 · 정렬 · 행높이
# ──────────────────────────────────────────────────────────────
def apply_borders(ws, theme, min_col, min_row, max_col, max_row):
    """표 내부 얇은 선 + 외곽 굵은 선.

    border_mode="rules"     : 격자 없음(헤더밑/소계/합계 가로선만). Ocean Blue.
    border_mode="statement" : 격자 없음 + 표 맨 위/맨 아래만 medium 굵은선. 정산표.
    border_mode="frame"     : 삼선표 — 좌/우 외곽선 없음, 상/하 굵게(medium), 내부 얇게(thin).
    기본 "grid"             : 외곽 medium + 내부 thin 전체 격자.
    """
    mode = theme.get("border_mode", "grid")
    if mode == "rules":
        return
    if mode == "statement":
        med = Side(style="medium", color=theme["border_color_outer"])
        for col in range(min_col, max_col + 1):
            tc = ws.cell(row=min_row, column=col); pt = tc.border
            tc.border = Border(left=pt.left, right=pt.right, top=med, bottom=pt.bottom)
            bc = ws.cell(row=max_row, column=col); pb = bc.border
            bc.border = Border(left=pb.left, right=pb.right, top=pb.top, bottom=med)
        return
    if mode == "frame":
        thin = Side(style="thin", color=theme["border_color"])
        thick = Side(style="medium", color=theme["border_color_outer"])
        for r in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                left = None if col == min_col else thin       # 좌/우 외곽선 없음
                right = None if col == max_col else thin
                top = thick if r == min_row else thin           # 상/하 굵게
                bottom = thick if r == max_row else thin
                ws.cell(row=r, column=col).border = Border(
                    left=left, right=right, top=top, bottom=bottom)
        return
    thin = Side(style="thin", color=theme["border_color"])
    thick = Side(style="medium", color=theme["border_color_outer"])
    for r in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            left = thick if col == min_col else thin
            right = thick if col == max_col else thin
            top = thick if r == min_row else thin
            bottom = thick if r == max_row else thin
            ws.cell(row=r, column=col).border = Border(
                left=left, right=right, top=top, bottom=bottom)


def style_body(ws, theme, header_row, min_col, min_row, max_col, max_row):
    """본문 폰트 · 세로 가운데정렬 · 행높이. 숫자는 오른쪽, 텍스트는 왼쪽 정렬."""
    body_font = Font(name=theme["font_name"], size=theme["font_size_body"],
                     color=theme.get("text_color"))
    body_start = max(min_row, header_row + 1)
    for r in range(body_start, max_row + 1):
        ws.row_dimensions[r].height = None    # 행높이 자동맞춤(고정값 미설정)
        for col in range(min_col, max_col + 1):
            c = ws.cell(row=r, column=col)
            c.font = body_font
            horiz = "right" if isinstance(c.value, (int, float)) else "left"
            c.alignment = Alignment(horizontal=horiz, vertical="center")


# ──────────────────────────────────────────────────────────────
# 3. 줄무늬 (zebra)
# ──────────────────────────────────────────────────────────────
def zebra_stripes(ws, theme, header_row, min_col, min_row, max_col, max_row):
    """헤더 아래 본문에서 짝수번째 데이터 행에 band_fill 적용."""
    fill = _fill(theme["band_fill"])
    body_start = max(min_row, header_row + 1)
    for i, r in enumerate(range(body_start, max_row + 1)):
        if i % 2 == 1:  # 0-based 두번째 행부터 번갈아
            for col in range(min_col, max_col + 1):
                ws.cell(row=r, column=col).fill = fill


# ──────────────────────────────────────────────────────────────
# 4. 틀 고정
# ──────────────────────────────────────────────────────────────
def freeze_below_header(ws, header_row, min_col):
    """헤더 바로 아래 + 첫 데이터 열 기준으로 틀 고정."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=min_col)


# ──────────────────────────────────────────────────────────────
# 5. 숫자 / 통화 서식
# ──────────────────────────────────────────────────────────────
def set_number_formats(ws, theme, col_formats, min_row, max_row, header_row):
    """열별 숫자 서식 적용.

    col_formats: {"C": "thousands", "D": "percent", 3: "currency_won"}
      값은 themes.number_formats 의 키이거나, 직접 number_format 문자열.
    """
    fmts = theme["number_formats"]
    body_start = max(min_row, header_row + 1)
    for col_key, fmt_name in col_formats.items():
        col = col_key if isinstance(col_key, int) else column_index_from_string(col_key)
        fmt = fmts.get(fmt_name, fmt_name)  # 키가 아니면 원문 그대로
        for r in range(body_start, max_row + 1):
            ws.cell(row=r, column=col).number_format = fmt


# ──────────────────────────────────────────────────────────────
# 6. 자동 열너비
# ──────────────────────────────────────────────────────────────
def autofit_columns(ws, min_col, min_row, max_col, max_row, min_width=8, max_width=60):
    """내용 길이 기반 열너비. 한글/전각은 폭 2로 계산."""
    for col in range(min_col, max_col + 1):
        longest = 0
        for r in range(min_row, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            width = sum(2 if ord(ch) > 0x1100 else 1 for ch in str(v))
            longest = max(longest, width)
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


# ──────────────────────────────────────────────────────────────
# 7. 조건부 서식 (음수 / 임계값 강조)
# ──────────────────────────────────────────────────────────────
def highlight_negatives(ws, theme, data_range):
    """data_range 안의 음수를 accent_color 글자색 + 굵게 강조."""
    font = Font(color=theme["accent_color"], bold=True)
    rule = CellIsRule(operator="lessThan", formula=["0"], font=font)
    ws.conditional_formatting.add(data_range, rule)


def highlight_threshold(ws, theme, data_range, operator, threshold):
    """임계값 조건부서식. operator: 'greaterThan','lessThan','equal' 등."""
    fill = _fill(theme["band_fill"])
    rule = CellIsRule(operator=operator, formula=[str(threshold)], fill=fill)
    ws.conditional_formatting.add(data_range, rule)


# ──────────────────────────────────────────────────────────────
# 8. 소계 / 합계 행 · 섹션 헤더 (정산표 관행)
# ──────────────────────────────────────────────────────────────
def style_total_row(ws, theme, row, min_col, max_col, fill=True):
    """소계/합계 행 강조.

    기본(grid) : medium 상단선 + thin 하단선 + 굵은 글씨 (+선택 배경).
    rules 모드 : 소계=상단 thin 선만, 합계(fill=True)=상단 thin + 하단 medium + 배경.
                 글자색은 total_font_color(보통 네이비). (Ocean Blue 관행)
    """
    rules = theme.get("border_mode", "grid") in ("rules", "statement")
    oc = theme["border_color_outer"]
    if rules:
        top = Side(style="thin", color=oc)
        bottom = Side(style="medium", color=oc) if fill else None
    else:
        top = Side(style="medium", color=oc)
        bottom = Side(style="thin", color=oc)
    bg = _fill(theme.get("total_fill", theme["band_fill"])) if fill else None
    fcolor = theme.get("total_font_color")
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name=theme["font_name"], size=theme["font_size_body"],
                      bold=True, color=fcolor)
        # 기존 좌우 테두리는 유지하고 상/하만 덮어쓴다
        prev = c.border
        c.border = Border(left=prev.left, right=prev.right,
                          top=top, bottom=(bottom if bottom is not None else prev.bottom))
        if bg is not None:
            c.fill = bg


def style_subheader_row(ws, theme, row, min_col, max_col):
    """섹션 구분 행: subheader_fill 배경 + 굵은 글씨.
    rules 모드에선 글자색을 total_font_color(네이비)로 맞춘다."""
    bg = _fill(theme.get("subheader_fill", theme["band_fill"]))
    fcolor = theme.get("total_font_color")
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = bg
        c.font = Font(name=theme["font_name"], size=theme["font_size_body"],
                      bold=True, color=fcolor)


def _fill_ranges(ws, hexcolor, ranges):
    bg = _fill(hexcolor)
    for rng in ranges:
        obj = ws[rng]
        if isinstance(obj, tuple):          # 범위 → 행 튜플들
            for rowcells in obj:
                for c in rowcells:
                    c.fill = bg
        else:                               # 단일 셀
            obj.fill = bg


def mark_cells(ws, theme, *, input=(), linked=()):
    """셀 종류를 색으로 표시 (DCF·감사조서 관행).

      input  : 직접 입력(하드코딩) 셀 → note_fill(#FFFBEF, 크림)
      linked : 타시트 연결 셀        → linked_fill(#CCECFF, 파랑)
      (수식 셀은 별도 표시 없음 = 검정 본문)

        mark_cells(ws, "default", input=["C5:C9", "E12"], linked=["C7"])
    """
    th = theme if isinstance(theme, dict) else get_theme(theme)
    if isinstance(input, str):
        input = [input]
    if isinstance(linked, str):
        linked = [linked]
    if input:
        _fill_ranges(ws, th.get("note_fill", "FFFBEF"), input)
    if linked:
        _fill_ranges(ws, th.get("linked_fill", "CCECFF"), linked)


def highlight_hardcoded(ws, theme, *ranges):
    """하드코딩(입력) 셀을 #FFFBEF 로 표시. (mark_cells(input=...) 의 별칭)"""
    th = theme if isinstance(theme, dict) else get_theme(theme)
    _fill_ranges(ws, th.get("note_fill", "FFFBEF"), ranges)


# ──────────────────────────────────────────────────────────────
# 워크북 기본폰트 — 값 없는 빈 셀까지 테마 폰트로 통일
# ──────────────────────────────────────────────────────────────
def set_workbook_font(ws, name, size=11):
    """워크북 Normal 스타일 폰트를 바꿔, 값이 없는 셀도 지정 폰트로 렌더되게 한다.
    (openpyxl 기본 minorFont=Calibri 때문에 빈 셀이 Calibri로 보이는 문제 해결)"""
    try:
        wb = ws.parent
        for st in wb._named_styles:
            if getattr(st, "name", "") == "Normal":
                st.font = Font(name=name, size=size)
                return
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────
# 통합 진입 함수
# ──────────────────────────────────────────────────────────────
def apply_theme(ws, theme=DEFAULT_THEME, header_row=1, data_range=None,
                title_cell=None, number_format_cols=None,
                zebra=False, freeze=False, autofit=True, highlight_neg=None,
                margin=True):
    """워크시트에 테마를 한 번에 적용한다.

    ws                : openpyxl Worksheet
    theme             : 테마 이름(str) 또는 토큰 dict
    header_row        : 헤더가 있는 행 번호 (1-based)
    data_range        : 표 영역 "A1:E20". None 이면 사용 영역 자동 감지.
    title_cell        : 시트 제목 셀 "B2" (선택). 제목 색은 데이터 열까지 확장된다.
    number_format_cols: {"C": "thousands", ...} 열별 숫자 서식 (선택)
    zebra/autofit/highlight_neg : 각 규칙 on/off
    freeze            : 틀 고정 (기본 규칙: 끔). 필요할 때만 True.
    margin            : A열·1행을 여백으로 비움 (기본 규칙: A열 너비 2.0, 1행 높이 ~27px).
    """
    th = theme if isinstance(theme, dict) else get_theme(theme)
    # 워크북 기본폰트(Normal)를 테마 폰트로 → 값 없는 빈 셀까지 동일 폰트(맑은 고딕 등)
    set_workbook_font(ws, th["font_name"], th.get("font_size_body", 11))
    min_col, min_row, max_col, max_row = _bounds(ws, data_range)

    # 음수 강조(굵게)는 테마 설정(excel_highlight_neg)을 따른다 — 미지정 시.
    # (False면 숫자서식 [Red]만 적용 → 기본 빨강·보통굵기)
    if highlight_neg is None:
        highlight_neg = th.get("excel_highlight_neg", False)

    # 기본 규칙: A열·1행을 동일한 작은 여백으로 비워둔다 (너비 2.0 / 높이 ~27px)
    if margin:
        ws.column_dimensions[get_column_letter(1)].width = 2.0
        ws.row_dimensions[1].height = 20.25   # ≈ 27px

    # 제목 (표 영역 밖일 수 있으므로 먼저)
    if title_cell:
        col_letter, trow = coordinate_from_string(title_cell)
        tcol = column_index_from_string(col_letter)
        tc = ws[title_cell]
        if th.get("title_fill"):
            # 제목 바 색(title_bg)을 데이터가 있는 마지막 열까지 가로로 확장
            bar = _fill(th.get("title_bg", th["header_bg"]))
            for col in range(tcol, max_col + 1):
                ws.cell(row=trow, column=col).fill = bar
            tc.font = Font(name=th["font_name"], size=th["font_size_title"],
                           bold=True,
                           color=th.get("title_font_color", th["header_font_color"]))
        else:
            tc.font = Font(name=th["font_name"], size=th["font_size_title"],
                           bold=True, color=th["title_color"])
        tc.alignment = Alignment(horizontal="left", vertical="center")

    # 본문 → 줄무늬 → 헤더 → 테두리 순서 (뒤가 앞을 덮어씀)
    style_body(ws, th, header_row, min_col, min_row, max_col, max_row)
    if zebra:
        zebra_stripes(ws, th, header_row, min_col, min_row, max_col, max_row)
    style_header_row(ws, th, header_row, min_col, max_col)
    apply_borders(ws, th, min_col, min_row, max_col, max_row)

    if number_format_cols:
        set_number_formats(ws, th, number_format_cols, min_row, max_row, header_row)
    if highlight_neg:
        body_range = "%s%d:%s%d" % (
            get_column_letter(min_col), header_row + 1,
            get_column_letter(max_col), max_row)
        highlight_negatives(ws, th, body_range)
    if freeze:
        freeze_below_header(ws, header_row, min_col)
    if autofit:
        autofit_columns(ws, min_col, min_row, max_col, max_row)

    return ws


# ──────────────────────────────────────────────────────────────
# 고수준 표 생성 (데이터 + 옵션 → 완성된 themed 표)
# ──────────────────────────────────────────────────────────────
def write_table(ws, headers, rows, theme=DEFAULT_THEME, *,
                title=None, number_cols=None, currency_cols=(), percent_cols=(),
                total_rows=(), subtotal_rows=(), section_rows=(),
                input_cells=(), linked_cells=(),
                title_cell="B2", header_row=4, start_col=2):
    """데이터를 받아 **한 번에** themed 표를 작성한다. 반환: 사용한 data_range.

    레이아웃: 제목=title_cell(기본 B2), 헤더=header_row(기본 4), 데이터=그 아래(B5~). A열 여백.
      headers       : 헤더 셀 리스트
      rows          : 2차원 리스트. None 셀은 비움(섹션 라벨 행 등).
      title         : B2 시트 제목 (None이면 제목 미작성)
      number_cols   : {열키: 서식키/문자열}. 열키=0-based 데이터열 인덱스 또는 'C' 같은 문자.
      currency_cols : 통화 열 인덱스들 → 테마 기본 통화서식(default=백만원, 그 외 accounting).
      percent_cols  : 퍼센트 열 인덱스들 → percent_acct.
      total/subtotal/section_rows : 데이터 1-based 인덱스(rows 기준).
      input_cells / linked_cells  : 입력=크림·연결=파랑 으로 표시할 범위/셀 (mark_cells).

    예) write_table(ws, ["계정","당기","전기"], data, theme="default",
                    title="재무상태표", currency_cols=[1,2], total_rows=[len(data)],
                    input_cells=["C5:D6"])
    """
    th = theme if isinstance(theme, dict) else get_theme(theme)
    n_cols = len(headers)

    if title is not None:
        ws[title_cell] = title
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=start_col + j, value=h)
    data_start = header_row + 1
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            if v is not None:
                ws.cell(row=data_start + i, column=start_col + j, value=v)
    last_row = data_start + len(rows) - 1
    data_range = "%s%d:%s%d" % (get_column_letter(start_col), header_row,
                                get_column_letter(start_col + n_cols - 1), last_row)

    # 숫자서식: currency/percent → number_cols 순으로 합치기 (열키는 문자로 정규화)
    def _col(key):
        return key if isinstance(key, str) else get_column_letter(start_col + key)
    fmts = {}
    cur_fmt = th.get("currency_format", "accounting")
    for k in currency_cols:
        fmts[_col(k)] = cur_fmt
    for k in percent_cols:
        fmts[_col(k)] = "percent_acct"
    for k, v in (number_cols or {}).items():
        fmts[_col(k)] = v

    apply_theme(ws, theme=th, header_row=header_row, data_range=data_range,
                title_cell=(title_cell if title is not None else None),
                number_format_cols=(fmts or None))

    def _srow(idx):
        return data_start + idx - 1
    mx = start_col + n_cols - 1
    for idx in section_rows:
        style_subheader_row(ws, th, _srow(idx), start_col, mx)
    for idx in subtotal_rows:
        style_total_row(ws, th, _srow(idx), start_col, mx, fill=False)
    for idx in total_rows:
        style_total_row(ws, th, _srow(idx), start_col, mx, fill=True)

    if input_cells or linked_cells:
        mark_cells(ws, th, input=list(input_cells), linked=list(linked_cells))

    return data_range
