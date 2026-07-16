# -*- coding: utf-8 -*-
"""
샘플 피드백 diff — 사용자가 엑셀에서 직접 고친 서식을 규칙 단위로 추출한다.

    python style_diff.py <샘플폴더>            보고서(feedback_report.md)만 생성
    python style_diff.py <샘플폴더> --apply    확정 제안을 excel_theme.py 에 자동 반영

동작:
  1. samples_<theme>.snapshot.json(생성 시점 스타일+역할맵) vs 편집된 xlsx 비교
  2. 셀 단위 차이를 (역할, 속성) 단위로 묶어 4가지로 분류:
     ① 자동 반영 가능   — 팔레트 토큰/숫자서식에 1:1 매핑되는 변경
     ② 코드 규칙 후보   — 토큰이 아니라 excel_theme.py 함수 로직 변경이 필요
     ③ 노이즈           — 열너비(autofit 산출물) 등 무시 대상
     ④ 모호/충돌        — 같은 역할인데 값이 갈리는 등 사용자 결정 필요
  3. --apply 시 ①만 반영 (palette_io.py 의 소스 수정 함수 재사용, .bak 백업)
"""
import glob
import json
import os
import re
import shutil
import sys
from collections import defaultdict

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)                       # excel-theme/
sys.path.insert(0, _ROOT)
sys.path.insert(0, _HERE)

from openpyxl import load_workbook

import palette_io
from excel_theme import NUMBER_FORMATS, PALETTES
from style_capture import (ATTR_LABELS, build_theme_resolver, capture_cell,
                           capture_dims, diff_styles)

SRC = os.path.join(_ROOT, "excel_theme.py")

ROLE_LABELS = {
    "title": "제목", "title_bar": "제목바", "group_header": "묶음헤더",
    "header": "헤더", "body": "본문", "body_zebra": "본문(줄무늬)",
    "section": "섹션행", "subtotal": "소계행", "total": "합계행",
    "input": "입력셀", "linked": "연결셀",
}

# (역할, 속성) → 팔레트 토큰 후보. ("colors", 키) 또는 ("palette", 키).
# 후보 중 '현재 팔레트 값 == 변경 전 값' 인 첫 토큰을 채택한다.
ROLE_ATTR_CANDIDATES = {
    ("header", "fill_rgb"):       [("colors", "excel_header_fill"), ("colors", "primary")],
    ("header", "font_color"):     [("colors", "excel_header_text"), ("colors", "header_font"),
                                   ("colors", "primary")],
    ("header", "font_size"):      [("palette", "excel_size_header")],
    ("header", "font_name"):      [("palette", "font_excel")],
    ("body", "font_name"):        [("palette", "font_excel")],
    ("body", "font_size"):        [("palette", "excel_size_body")],
    ("body", "font_color"):       [("colors", "text")],
    ("body_zebra", "fill_rgb"):   [("colors", "band")],
    ("body_zebra", "font_name"):  [("palette", "font_excel")],
    ("body_zebra", "font_size"):  [("palette", "excel_size_body")],
    ("body_zebra", "font_color"): [("colors", "text")],
    ("title", "font_size"):       [("palette", "excel_title_size")],
    ("title", "font_name"):       [("palette", "font_excel")],
    ("title", "fill_rgb"):        [("colors", "title_bg"), ("colors", "primary")],
    ("title_bar", "fill_rgb"):    [("colors", "title_bg"), ("colors", "primary")],
    ("title", "font_color"):      [("colors", "title_font_color"), ("colors", "title_color"),
                                   ("colors", "header_font")],
    ("group_header", "fill_rgb"): [("colors", "subheader")],
    ("group_header", "font_color"): [("colors", "primary_dark")],
    ("section", "fill_rgb"):      [("colors", "subheader")],
    ("section", "font_color"):    [("colors", "primary_dark")],
    ("subtotal", "font_color"):   [("colors", "primary_dark")],
    ("total", "fill_rgb"):        [("colors", "band")],
    ("total", "font_color"):      [("colors", "primary_dark")],
    ("input", "fill_rgb"):        [("colors", "note")],
    ("linked", "fill_rgb"):       [("colors", "linked")],
}

# 코드 규칙 후보일 때 안내할 excel_theme.py 함수
ROLE_FUNC = {
    "header": "style_header_row", "body": "style_body", "body_zebra": "style_body/zebra_stripes",
    "section": "style_subheader_row", "group_header": "style_subheader_row",
    "subtotal": "style_total_row", "total": "style_total_row",
    "title": "apply_theme(제목 처리)", "title_bar": "apply_theme(제목 처리)",
    "input": "mark_cells", "linked": "mark_cells",
}


def _fmt_val(attr, v):
    if v is None:
        return "(없음)"
    if attr.startswith("border_"):
        style, _, color = str(v).partition("|")
        return f"{style}(#{color})"
    if attr in ("fill_rgb", "font_color"):
        return f"#{v}"
    return str(v)


# ──────────────────────────────────────────────────────────────
# 1) 수집: 스냅샷 vs 편집본
# ──────────────────────────────────────────────────────────────
def collect_diffs(theme, xlsx_path, snapshot):
    """셀/서식키/치수 차이 수집."""
    wb = load_workbook(xlsx_path)
    resolver = build_theme_resolver(wb)
    cell_diffs, fmt_diffs, dim_diffs = [], [], []

    for sheet_name, meta in snapshot["sheets"].items():
        if sheet_name not in wb.sheetnames:
            cell_diffs.append((sheet_name, "(시트)", "sheet", "missing", "있음", "삭제됨"))
            continue
        ws = wb[sheet_name]
        roles, fmt_keys = meta["roles"], meta.get("fmt_keys", {})

        for coord, old in meta["styles"].items():
            new = capture_cell(ws[coord], resolver)
            changed = diff_styles(old, new)
            role = roles.get(coord, "body")
            for attr, (ov, nv) in changed.items():
                if attr == "number_format" and coord in fmt_keys:
                    fmt_diffs.append((sheet_name, coord, fmt_keys[coord], ov, nv))
                else:
                    cell_diffs.append((sheet_name, coord, role, attr, ov, nv))

        # 행높이/열너비 (비교 대상은 스냅샷에 기록된 키에서 파생)
        old_dims = meta.get("dims", {})
        new_dims = capture_dims(ws, cols=list(old_dims.get("col_widths", {})),
                                rows=[int(r) for r in old_dims.get("row_heights", {})])
        row_roles = _row_roles(roles)
        # 우리가 설정하지 않은(스냅샷=None) 치수는 Excel 이 저장 시 자동값을
        # 채워 넣으므로 비교하지 않는다. 미세 재계산(20→20.05)은 허용오차로 흡수.
        for r, ov in old_dims.get("row_heights", {}).items():
            if ov is None:
                continue
            nv = new_dims["row_heights"].get(r)
            if not _num_eq(ov, nv, tol=0.3):
                dim_diffs.append(("row", sheet_name, r, row_roles.get(int(r), "기타"), ov, nv))
        for letter, ov in old_dims.get("col_widths", {}).items():
            if ov is None:
                continue
            nv = new_dims["col_widths"].get(letter)
            if not _num_eq(ov, nv, tol=0.5):
                dim_diffs.append(("col", sheet_name, letter, None, ov, nv))
    return cell_diffs, fmt_diffs, dim_diffs


def _num_eq(a, b, tol=0.05):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tol


def _row_roles(roles):
    """행번호 → 'header'/'body'/기타 (행높이 토큰 매핑용)."""
    out = {}
    for coord, role in roles.items():
        m = re.match(r"[A-Z]+(\d+)$", coord)
        if not m:
            continue
        r = int(m.group(1))
        if role == "header":
            out[r] = "header"
        elif role in ("body", "body_zebra", "section", "subtotal", "total",
                      "input", "linked") and out.get(r) != "header":
            out[r] = "body"
        elif role in ("title", "title_bar") and r not in out:
            out[r] = "title"
        elif r not in out:
            out[r] = "기타"
    return out


# ──────────────────────────────────────────────────────────────
# 2) 집계: (테마, 역할, 속성) 단위 분류
# ──────────────────────────────────────────────────────────────
def aggregate(per_theme):
    """per_theme: {theme: (cell_diffs, fmt_diffs, dim_diffs, role_counts)}"""
    confirmed, code_rules, noise, ambiguous = [], [], [], []
    token_targets = defaultdict(list)   # (theme, kind, token) → 제안 인덱스 (충돌 검사)

    for theme, (cell_diffs, fmt_diffs, dim_diffs, role_counts) in per_theme.items():
        pal = PALETTES[theme]

        groups = defaultdict(list)
        for sheet, coord, role, attr, ov, nv in cell_diffs:
            groups[(role, attr)].append((sheet, coord, ov, nv))

        for (role, attr), items in sorted(groups.items()):
            olds = {str(ov) for _, _, ov, _ in items}
            news = {str(nv) for _, _, _, nv in items}
            cells = _cells_summary(items)
            base = {"theme": theme, "role": role, "attr": attr, "cells": cells,
                    "count": len(items), "total": role_counts.get(role, 0)}
            if len(news) > 1 or len(olds) > 1:
                ambiguous.append({**base, "detail": _ambiguous_detail(attr, items)})
                continue
            ov, nv = items[0][2], items[0][3]
            token = _match_token(pal, role, attr, ov)
            if token:
                kind, key = token
                p = {**base, "kind": kind, "token": key, "old": ov, "new": nv}
                confirmed.append(p)
                token_targets[(theme, kind, key)].append(p)
            else:
                code_rules.append({**base, "old": ov, "new": nv,
                                   "func": ROLE_FUNC.get(role, "excel_theme.py")})

        # 숫자서식 (NUMBER_FORMATS 는 테마 공통)
        fmt_groups = defaultdict(list)
        for sheet, coord, key, ov, nv in fmt_diffs:
            fmt_groups[key].append((sheet, coord, ov, nv))
        for key, items in sorted(fmt_groups.items()):
            news = {nv for _, _, _, nv in items}
            cells = _cells_summary(items)
            base = {"theme": theme, "role": f"숫자서식 '{key}'", "attr": "number_format",
                    "cells": cells, "count": len(items), "total": 0}
            if len(news) > 1:
                ambiguous.append({**base, "detail": _ambiguous_detail("number_format", items)})
                continue
            nv = items[0][3]
            p = {**base, "kind": "nf", "token": key,
                 "old": NUMBER_FORMATS.get(key, "?"), "new": nv}
            confirmed.append(p)
            token_targets[("(공통)", "nf", key)].append(p)

        # 행높이/열너비
        row_groups = defaultdict(list)
        for kind, sheet, key, row_role, ov, nv in dim_diffs:
            if kind == "col":
                noise.append({"theme": theme, "what": f"열너비 {sheet}!{key}",
                              "old": ov, "new": nv})
            else:
                row_groups[row_role].append((sheet, key, ov, nv))
        for row_role, items in sorted(row_groups.items()):
            news = {nv for _, _, _, nv in items}
            cells = ", ".join(f"{s}!{k}행" for s, k, _, _ in items[:4]) + \
                    (f" 외 {len(items)-4}곳" if len(items) > 4 else "")
            base = {"theme": theme, "role": f"{row_role} 행높이", "attr": "row_height",
                    "cells": cells, "count": len(items), "total": 0}
            if len(news) > 1:
                ambiguous.append({**base, "detail": _ambiguous_detail("row_height", items)})
            elif row_role in ("header", "body"):
                key = "excel_row_header" if row_role == "header" else "excel_row_body"
                nv = items[0][3]
                p = {**base, "kind": "palette", "token": key,
                     "old": pal.get(key), "new": nv}
                confirmed.append(p)
                token_targets[(theme, "palette", key)].append(p)
            else:
                code_rules.append({**base, "old": items[0][2], "new": items[0][3],
                                   "func": "apply_theme(여백/제목 행높이)"})

    # 같은 토큰에 서로 다른 값 → 충돌 = 모호로 강등
    for (theme, kind, key), plist in token_targets.items():
        if len({str(p["new"]) for p in plist}) > 1:
            for p in plist:
                confirmed.remove(p)
                ambiguous.append({**p, "detail": f"토큰 {key} 에 서로 다른 값 제안(충돌)"})
    return confirmed, code_rules, noise, ambiguous


def _cells_summary(items, limit=4):
    cells = [f"{s}!{c}" for s, c, _, _ in items]
    return ", ".join(cells[:limit]) + (f" 외 {len(cells)-limit}곳" if len(cells) > limit else "")


def _ambiguous_detail(attr, items):
    return "; ".join(f"{s}!{c}: {_fmt_val(attr, ov)}→{_fmt_val(attr, nv)}"
                     for s, c, ov, nv in items[:6]) + (" ..." if len(items) > 6 else "")


def _match_token(pal, role, attr, old_val):
    """변경 전 값과 일치하는 팔레트 토큰을 찾는다. 없으면 None(=코드 규칙 후보)."""
    for kind, key in ROLE_ATTR_CANDIDATES.get((role, attr), []):
        if kind == "colors":
            cur = pal["colors"].get(key)
            if cur and old_val and str(cur).upper() == str(old_val).upper():
                return (kind, key)
        else:
            cur = pal.get(key)
            if cur is None:
                continue
            if attr in ("font_size",) or key.startswith("excel_row"):
                try:
                    if abs(float(cur) - float(old_val)) < 0.3:
                        return (kind, key)
                except (TypeError, ValueError):
                    continue
            elif str(cur) == str(old_val):
                return (kind, key)
    return None


# ──────────────────────────────────────────────────────────────
# 3) 보고서
# ──────────────────────────────────────────────────────────────
def render_report(out_path, confirmed, code_rules, noise, ambiguous, sources):
    L = ["# 샘플 피드백 보고서", ""]
    L.append("비교 대상: " + ", ".join(os.path.basename(s) for s in sources))
    L.append("")
    total = len(confirmed) + len(code_rules) + len(ambiguous)
    if total == 0:
        L.append("**변경 없음** — 샘플에서 서식 변경이 감지되지 않았습니다.")
        if noise:
            L.append(f"(노이즈로 무시한 항목 {len(noise)}건: 열너비 등)")
    else:
        L.append(f"감지된 규칙 변경: 자동반영 {len(confirmed)} / 코드규칙 {len(code_rules)}"
                 f" / 모호 {len(ambiguous)} / 노이즈 {len(noise)}")
    L.append("")

    L.append("## ① 자동 반영 가능 (테마 토큰) — `dev.bat -Diff -Apply` 로 반영")
    L.append("")
    if confirmed:
        L.append("| 테마 | 대상 | 토큰 | 변경 전 → 후 | 바뀐 셀 (역할 내 비율) |")
        L.append("|---|---|---|---|---|")
        for p in confirmed:
            role_kr = ROLE_LABELS.get(p["role"], p["role"])
            attr_kr = ATTR_LABELS.get(p["attr"], p["attr"])
            cover = f"{p['count']}/{p['total']}" if p["total"] else f"{p['count']}곳"
            L.append(f"| {p['theme']} | {role_kr} {attr_kr} | `{p['token']}` "
                     f"| {_fmt_val(p['attr'], p['old'])} → {_fmt_val(p['attr'], p['new'])} "
                     f"| {p['cells']} ({cover}) |")
    else:
        L.append("(없음)")
    L.append("")

    L.append("## ② 코드 규칙 후보 — 토큰이 아니라 함수 로직 변경 필요")
    L.append("")
    if code_rules:
        for p in code_rules:
            role_kr = ROLE_LABELS.get(p["role"], p["role"])
            attr_kr = ATTR_LABELS.get(p["attr"], p["attr"])
            L.append(f"- [{p['theme']}] {role_kr} {attr_kr}: "
                     f"{_fmt_val(p['attr'], p['old'])} → {_fmt_val(p['attr'], p['new'])} "
                     f"({p['cells']})")
            L.append(f"  - Claude에게 이렇게 요청하세요: "
                     f"\"excel_theme.py 의 `{p['func']}` 에서 {p['theme']} 테마 "
                     f"{role_kr} {attr_kr}를 {_fmt_val(p['attr'], p['new'])} 로 바꿔줘\"")
    else:
        L.append("(없음)")
    L.append("")

    L.append("## ③ 노이즈로 간주 (반영 안 함)")
    L.append("")
    if noise:
        L.append(f"- 열너비 변경 {len(noise)}건 — 열너비는 autofit(내용 길이)으로 계산되어 "
                 "테마 규칙이 아닙니다. 고정폭 정책으로 바꾸려면 Claude에게 요청하세요.")
        for n in noise[:8]:
            L.append(f"  - [{n['theme']}] {n['what']}: {n['old']} → {n['new']}")
        if len(noise) > 8:
            L.append(f"  - ... 외 {len(noise) - 8}건")
    else:
        L.append("(없음)")
    L.append("")

    L.append("## ④ 모호 / 충돌 — 값이 갈려서 자동 반영하지 않음")
    L.append("")
    if ambiguous:
        for p in ambiguous:
            role_kr = ROLE_LABELS.get(p["role"], p["role"])
            attr_kr = ATTR_LABELS.get(p["attr"], p["attr"])
            L.append(f"- [{p['theme']}] {role_kr} {attr_kr} ({p['cells']})")
            L.append(f"  - {p.get('detail', '')}")
        L.append("")
        L.append("원하는 쪽을 정해 Claude에게 알려주거나, 샘플에서 한 가지 값으로 통일 후 "
                 "다시 -Diff 하세요.")
    else:
        L.append("(없음)")
    L.append("")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


# ──────────────────────────────────────────────────────────────
# 4) 자동 반영 (palette_io 의 소스 수정 함수 재사용)
# ──────────────────────────────────────────────────────────────
def apply_confirmed(confirmed):
    if not confirmed:
        print("자동 반영할 항목이 없습니다.")
        return False
    with open(SRC, encoding="utf-8") as f:
        src = f.read()

    per_theme = defaultdict(lambda: {"colors": {}, "fonts": {}, "sizes": {}})
    nf_updates = {}
    for p in confirmed:
        if p["kind"] == "nf":
            nf_updates[p["token"]] = p["new"]
        elif p["kind"] == "colors":
            per_theme[p["theme"]]["colors"][p["token"]] = str(p["new"]).upper()
        elif p["token"].startswith("font"):
            per_theme[p["theme"]]["fonts"][p["token"]] = str(p["new"])
        else:
            per_theme[p["theme"]]["sizes"][p["token"]] = int(round(float(p["new"])))

    total = []
    for theme, upd in per_theme.items():
        span = palette_io._theme_block_span(src, theme)
        if not span:
            print(f"  (건너뜀: 소스에서 '{theme}' 블록을 못 찾음)")
            continue
        s, e = span
        new_block, changed = palette_io._apply_updates(
            src[s:e], upd["colors"], upd["fonts"], {}, upd["sizes"])
        if changed:
            src = src[:s] + new_block + src[e:]
            total.append((theme, changed))
    for key, val in nf_updates.items():
        src, n = palette_io._replace_nf(src, key, val)
        if n:
            total.append(("숫자서식", [key]))

    if not total:
        print("반영된 변경 없음 (소스에서 대상 토큰을 못 찾음).")
        return False
    shutil.copyfile(SRC, SRC + ".bak")
    with open(SRC, "w", encoding="utf-8") as f:
        f.write(src)
    print("excel_theme.py 갱신 완료 (백업: excel_theme.py.bak)")
    for name, ch in total:
        print(f"  [{name}] " + ", ".join(str(c) for c in ch))
    return True


# ──────────────────────────────────────────────────────────────
def main(sample_dir, do_apply):
    snaps = sorted(glob.glob(os.path.join(sample_dir, "samples_*.snapshot.json")))
    if not snaps:
        print(f"스냅샷이 없습니다: {sample_dir}\\samples_*.snapshot.json")
        print("먼저  dev.bat -Samples  로 샘플을 생성하세요.")
        return 1

    per_theme, sources = {}, []
    for snap_path in snaps:
        with open(snap_path, encoding="utf-8") as f:
            snapshot = json.load(f)
        theme = snapshot["theme"]
        if theme not in PALETTES:
            print(f"  (건너뜀: 알 수 없는 테마 '{theme}')")
            continue
        xlsx_path = os.path.join(sample_dir, snapshot["file"])
        if not os.path.exists(xlsx_path):
            print(f"  (건너뜀: 편집본 없음 {xlsx_path})")
            continue
        sources.append(xlsx_path)
        cell_diffs, fmt_diffs, dim_diffs = collect_diffs(theme, xlsx_path, snapshot)
        role_counts = defaultdict(int)
        for m in snapshot["sheets"].values():
            for role in m["roles"].values():
                role_counts[role] += 1
        per_theme[theme] = (cell_diffs, fmt_diffs, dim_diffs, dict(role_counts))

    confirmed, code_rules, noise, ambiguous = aggregate(per_theme)
    report_path = os.path.join(sample_dir, "feedback_report.md")
    render_report(report_path, confirmed, code_rules, noise, ambiguous, sources)

    n_changes = len(confirmed) + len(code_rules) + len(ambiguous)
    if n_changes == 0:
        print("변경 없음 — 샘플 서식이 스냅샷과 일치합니다."
              + (f" (노이즈 {len(noise)}건 무시)" if noise else ""))
    else:
        print(f"감지: 자동반영 {len(confirmed)} / 코드규칙 {len(code_rules)} / "
              f"모호 {len(ambiguous)} / 노이즈 {len(noise)}")
    print(f"보고서: {report_path}")

    if do_apply:
        applied = apply_confirmed(confirmed)
        return 0 if (applied or not confirmed) else 1
    if confirmed:
        print("반영하려면:  dev.bat -Diff -Apply")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("사용법: python style_diff.py <샘플폴더> [--apply]")
        sys.exit(2)
    sys.exit(main(args[0], "--apply" in sys.argv))
